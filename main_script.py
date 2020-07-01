import numpy as np
import pandas as pd

from own_package.features_labels import read_excel_data, read_excel_dataloader, Fl_master, Fl_pca, Fl_ar, \
    Fl_cw, hparam_selection
from own_package.boosting import ComponentwiseL2Boost, ComponentwiseL2BoostDropout
from own_package.others import create_results_directory
from own_package.pre_processing import type_transformations
import openpyxl, time
from openpyxl.utils.dataframe import dataframe_to_rows
from own_package.others import print_array_to_excel


def selector(case, excel_dir=None, var_name=None):
    if case == 0:
        excel_dir = './excel/dataset_blanks.xlsx'
        type_transformations(excel_dir=excel_dir,
                             y_selection=['W875RX1', 'DPCERA3M086SBEA', 'CMRMTSPLx', 'INDPRO',
                                          'PAYEMS', 'WPSFD49207', 'CPIAUCSL', 'CPIULFSL'],
                             h_steps=[1, 3, 6, 12, 24])
    elif case == 1:
        # Testing filling out missing observation using iterated EM method
        excel_dir = './excel/dataset_filled.xlsx'
        features, labels, time_stamp, features_names, labels_names, label_type = read_excel_data(excel_dir=excel_dir)
        fl_master = Fl_master(x=features, yo=labels, time_stamp=time_stamp,
                              features_names=features_names, labels_names=labels_names)
        fl_master.iterated_em(features=features, labels=labels, pca_p=9, max_iter=10000, tol=0.1, excel_dir=excel_dir)
    elif case == 2:
        # Testing pca k selection using IC criteria
        excel_dir = './excel/dataset_filled.xlsx'
        features, labels, time_stamp, features_names, labels_names, label_type = read_excel_data(excel_dir=excel_dir)
        fl_master = Fl_master(x=features, yo=labels, time_stamp=time_stamp,
                              features_names=features_names, labels_names=labels_names)
        (f_tv, f_tt), (yo_tv, yo_t), (y_tv, y_tt), \
        (ts_tv, ts_tt), (tidx_tv, tidx_tt), (nobs_tv, nobs_tt) = fl_master.percentage_split(0)
        fl = Fl_pca(val_split=0.2, x=f_tv, yo=yo_tv, y=y_tv,
                    time_stamp=ts_tv, time_idx=tidx_tv, features_names=features_names, labels_names=labels_names)
        r, ic_store = fl.pca_k_selection(lower_k=5, upper_k=40)
        print(ic_store)
    elif case == 3:
        # Training and validation set
        var_name = var_name
        excel_dir = excel_dir
        results_dir = create_results_directory('./results/test')
        output = read_excel_dataloader(excel_dir=excel_dir)
        fl_master = Fl_master(x=output[0], features_names=output[1],
                              yo=output[2], labels_names=output[3],
                              y=output[4], y_names=output[5],
                              time_stamp=output[6])
        (f_tv, f_tt), (yo_tv, yo_t), (y_tv, y_tt), \
        (ts_tv, ts_tt), (tidx_tv, tidx_tt), (nobs_tv, nobs_tt) = fl_master.percentage_split(0.2)

        fl_pca = Fl_pca(val_split=0.2, x=f_tv, yo=yo_tv, y=y_tv,
                        time_stamp=ts_tv, time_idx=tidx_tv,
                        features_names=fl_master.features_names, labels_names=fl_master.labels_names,
                        y_names=fl_master.y_names)

        fl_ar = Fl_ar(val_split=0.2, x=f_tv, yo=yo_tv, y=y_tv,
                      time_stamp=ts_tv, time_idx=tidx_tv,
                      features_names=fl_master.features_names, labels_names=fl_master.labels_names,
                      y_names=fl_master.y_names)
        h_steps = [1, 3, 6, 12, 24]
        type_store = ['PLS', 'PLS', 'AIC_BIC', 'AIC_BIC', 'PLS', 'AIC_BIC']
        model_store = ['AR', 'PCA', 'AR', 'PCA', 'UMAP', 'UMAP']
        # type_store = ['PLS']
        # model_store = ['UMAP']
        for type, model in zip(type_store, model_store):
            print('{}_{} Experiment'.format(type, model))
            wb = openpyxl.Workbook()
            if model == 'AR':
                bounds_m = [1, 1]
                bounds_p = [1, 12]
                fl = fl_ar
            elif model == 'PCA':
                bounds_m = [1, 9]
                bounds_p = [1, 12]
                fl = fl_pca
            elif model == 'UMAP':
                bounds_m = [1, 3]
                bounds_p = [1, 12]
                fl = fl_pca
            for idx, h in enumerate(h_steps):
                start = time.time()
                df = hparam_selection(self=fl, model=model, type=type, bounds_m=bounds_m, bounds_p=bounds_p,
                                      h=h, h_idx=idx, h_max=max(h_steps), r=9, results_dir=results_dir,
                                      extension=False)
                wb.create_sheet('{}_h_{}'.format(type, h))
                sheet_name = wb.sheetnames[-1]
                ws = wb[sheet_name]

                print_array_to_excel(array=['h = {}'.format(h), 'r = 9', var_name], first_cell=(1, 1), ws=ws, axis=1)
                rows = dataframe_to_rows(df)

                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx + 1, column=c_idx, value=value)
                end = time.time()
                print('h = {} completed. Time Taken = {}'.format(h, end - start))

            wb.save(filename='{}/{}_{}_{}.xlsx'.format(results_dir, var_name, model, type))
    elif case == 4:
        # Test set
        var_name = var_name
        excel_dir = excel_dir
        results_dir = create_results_directory('./results/{}'.format(var_name))
        output = read_excel_dataloader(excel_dir=excel_dir)
        fl_master = Fl_master(x=output[0], features_names=output[1],
                              yo=output[2], labels_names=output[3],
                              y=output[4], y_names=output[5],
                              time_stamp=output[6])
        (f_tv, f_tt), (yo_tv, yo_t), (y_tv, y_tt), \
        (ts_tv, ts_tt), (tidx_tv, tidx_tt), (nobs_tv, nobs_tt) = fl_master.percentage_split(0)
        #f_tv = np.array([list(range(1,21))]*5).T
        #yo_tv = np.array([list(range(1,21))]).T
        #y_tv = np.array([list(range(1,21))]*5).T
        fl_pca = Fl_pca(val_split=0.2, x=f_tv, yo=yo_tv, y=y_tv,
                        time_stamp=ts_tv, time_idx=tidx_tv,
                        features_names=fl_master.features_names, labels_names=fl_master.labels_names,
                        y_names=fl_master.y_names)
        fl_ar = Fl_ar(val_split=0.2, x=f_tv, yo=yo_tv, y=y_tv,
                      time_stamp=ts_tv, time_idx=tidx_tv,
                      features_names=fl_master.features_names, labels_names=fl_master.labels_names,
                      y_names=fl_master.y_names)
        fl_cw = Fl_cw(val_split=0.2, x=f_tv, yo=yo_tv, y=y_tv,
                        time_stamp=ts_tv, time_idx=tidx_tv,
                        features_names=fl_master.features_names, labels_names=fl_master.labels_names,
                        y_names=fl_master.y_names)
        #h_steps = [6]
        #h_idx_store = [2]  # h = 1 corresponds to h_idx=0, h=3 is h_idx=1, and so on
        h_steps = [1,3,6,12,24]
        h_idx_store = [0,1,2,3,4]
        #type_store = ['PLS', 'PLS', 'PLS']
        #model_store = ['CW1']
        model_store = ['CWd3','CWd4','CWd7','CWd8']
        #model_store =['CWd5','CWd6']
        type_store = ['PLS'] * len(model_store)
        # type_store = ['AIC_BIC']
        # model_store = ['UMAP']
        for type, model in zip(type_store, model_store):
            print('{}_{} Experiment'.format(type, model))
            wb = openpyxl.Workbook()
            if model == 'AR':
                bounds_m = [1, 1]
                bounds_p = [1, 12]
                fl = fl_ar
                kwargs = {}
            elif model == 'PCA':
                bounds_m = [1, 9]
                bounds_p = [1, 12]
                fl = fl_pca
                kwargs = {}
            elif model == 'UMAP':
                bounds_m = [1, 3]
                bounds_p = [1, 12]
                fl = fl_pca
                kwargs = {}
            elif model[:-1] == 'CW':
                bounds_m = [3,3]
                bounds_p = [12,12]
                fl = fl_cw
                kwargs = {'cw_model_class':ComponentwiseL2Boost,
                          'cw_hparams': {'m_max': 1000, 'learning_rate': 0.3, 'ic_mode': 'aic', 'dropout': 0.5}}
            elif model[:-1] == 'CWd':
                bounds_m = [12,12]
                bounds_p = [24,24]
                fl = fl_cw
                kwargs = {'cw_model_class': ComponentwiseL2BoostDropout,
                          'cw_hparams': {'m_max': 250, 'learning_rate': 0.3, 'ic_mode': 'aic', 'dropout': 0.5}}
            else:
                raise KeyError('Invalid model mode selected.')

            for h_idx, h in zip(h_idx_store, h_steps):
                start = time.time()
                df = hparam_selection(fl=fl, model=model, type=type, bounds_m=bounds_m, bounds_p=bounds_p, h=h,
                                      h_idx=h_idx,
                                      h_max=max(h_steps), r=8, results_dir=results_dir,
                                      extension=False, rolling=True, **kwargs)
                wb.create_sheet('{}_h_{}'.format(type, h))
                sheet_name = wb.sheetnames[-1]
                ws = wb[sheet_name]

                print_array_to_excel(array=['h = {}'.format(h), 'r = 9', var_name], first_cell=(1, 1), ws=ws, axis=1)
                rows = dataframe_to_rows(df)

                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        ws.cell(row=r_idx + 1, column=c_idx, value=value)
                end = time.time()
                print('h = {} completed. Time Taken = {}'.format(h, end - start))

            wb.save(filename='{}/{}_{}_{}.xlsx'.format(results_dir, var_name, model, type))

    pass

if __name__ == '__main__':
    # selector(0)
    #selector(1)
    selector(4, excel_dir='./excel/dataset2/CPIAUCSL_data_loader.xlsx', var_name='testset_CPIAm12p24')
    # selector(3, excel_dir='./excel/DPC_data_loader.xlsx', var_name='DPC')

    # selector(3, excel_dir='./excel/WPSFD49207_data_loader.xlsx', var_name='WPSFD49207')
    # selector(4, excel_dir='./excel/WPSFD49207_data_loader.xlsx', var_name='WPSFD49207')

    # selector(3, excel_dir='./excel/IND_data_loader.xlsx', var_name='IND')
    # selector(4, excel_dir='./excel/IND_data_loader.xlsx', var_name='IND')

    # selector(3, excel_dir='./excel/PAY_data_loader.xlsx', var_name='PAY')
    # selector(4, excel_dir='./excel/PAY_data_loader.xlsx', var_name='PAY')
