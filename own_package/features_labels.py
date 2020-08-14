import numpy as np
import pandas as pd
import copy, math, pickle
import time, itertools
import concurrent.futures
import matplotlib.pyplot as plt
import xgboost as xgb
from sklearn.preprocessing import StandardScaler
from scipy.linalg import eigh
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import umap, shap
import statsmodels.api as sm
from sklearn.model_selection import cross_val_score
from skopt import gp_minimize
from skopt.space import Real, Integer
from skopt.utils import use_named_args

from own_package.others import create_results_directory, print_df_to_excel
from own_package.boosting import Xgboost


def read_excel_data(excel_dir):
    features = pd.read_excel(excel_dir, sheet_name='features')
    features_names = features.columns.values.tolist()
    features = features.values

    labels = pd.read_excel(excel_dir, sheet_name='labels')
    labels_names = [labels.columns.values.tolist()[0]]
    try:
        label_type = labels.values[1, 0]
    except:
        label_type = None
    labels = labels.values[:, 0][..., None].astype(np.float)

    time_stamp = np.copy(features[:, 0])
    features = np.copy(features[:, 1:]).astype(np.float)

    return features, labels, time_stamp, features_names, labels_names, label_type


def read_excel_dataloader(excel_dir):
    xls = pd.ExcelFile(excel_dir)
    df = pd.read_excel(xls, 'x', index_col=0)
    x_names = df.columns.values
    x = df.values

    df = pd.read_excel(xls, 'yo', index_col=0)
    yo_names = df.columns.values
    yo = df.values

    df = pd.read_excel(xls, 'y', index_col=0)
    y_names = df.columns.values
    y = df.values

    time_stamp = df.index.values

    return (x, x_names, yo, yo_names, y, y_names, time_stamp)


def hparam_selection(fl, model, type, bounds_m, bounds_p, h, h_idx, h_max, r, results_dir, extension=False,
                     rolling=False, **kwargs):
    if extension:
        m_init = list(range(bounds_m[0], bounds_m[1] + 1))
        p_init = list(range(bounds_p[0], bounds_p[1] + 1))

        h_param_init = list(itertools.product(m_init, p_init))

        m_cancel = list(range(1, 3 + 1))
        p_cancel = list(range(1, 6 + 1))

        h_param_cancel = list(itertools.product(m_cancel, p_cancel))

        hparams_store = [x for x in h_param_init if x not in h_param_cancel]
    else:
        m_store = list(range(bounds_m[0], bounds_m[1] + 1))
        p_store = list(range(bounds_p[0], bounds_p[1] + 1))
        hparams_store = list(itertools.product(m_store, p_store))

    rmse_store = []
    aic_t_store = []
    bic_t_store = []
    y_hat_store = []

    m_max = bounds_m[1]
    p_max = bounds_p[1]

    if model == 'PCA' or model == 'UMAP':
        if model == 'PCA':
            factor_model = fl.pca_factor_estimation
        else:
            factor_model = fl.umap_factor_estimation

        if type == 'PLS':
            data_store_save_dir = create_results_directory('{}/{}_h{}'.format(results_dir, model, h))
            for m, p in hparams_store:
                y_hat, _, rmse, results_t = fl.pls_expanding_window(h=h, m=m, p=p, r=r,
                                                                    factor_model=factor_model,
                                                                    x_t=fl.x_t, yo_t=fl.yo_t,
                                                                    y_t=fl.y_t[:, h_idx][..., None],
                                                                    x_v=fl.x_v, yo_v=fl.yo_v,
                                                                    y_v=fl.y_v[:, h_idx][..., None],
                                                                    rolling=rolling,
                                                                    save_dir=data_store_save_dir,
                                                                    save_name=model,
                                                                    )

                rmse_store.append(rmse)
                aic_t_store.append(results_t.aic)
                bic_t_store.append(results_t.bic)
                y_hat_store.append(y_hat)
            df = pd.DataFrame(data=np.concatenate((np.array(hparams_store),
                                                   np.array(rmse_store)[..., None],
                                                   np.array(aic_t_store)[..., None],
                                                   np.array(bic_t_store)[..., None],
                                                   np.array(y_hat_store)), axis=1),
                              columns=['m', 'p', 'Val RMSE', 'AIC_t', 'BIC_t'] + fl.y_v[:,
                                                                                 h_idx].flatten().tolist())


        elif type == 'AIC_BIC':
            for m, p in hparams_store:
                aic, bic = fl.aic_bic_for_pca_umap(h=h, m=m, p=p, r=r, h_max=h_max, m_max=m_max, p_max=p_max,
                                                   factor_model=factor_model,
                                                   x=fl.x, yo=fl.yo, y=fl.y[:, h_idx][..., None])
                rmse_store.append(-1)
                aic_t_store.append(aic)
                bic_t_store.append(bic)
            df = pd.DataFrame(data=np.concatenate((np.array(hparams_store),
                                                   np.array(rmse_store)[..., None],
                                                   np.array(aic_t_store)[..., None],
                                                   np.array(bic_t_store)[..., None]), axis=1),
                              columns=['m', 'p', 'Val RMSE', 'AIC_t', 'BIC_t'])

    elif model == 'AR':
        if type == 'PLS':
            data_store_save_dir = create_results_directory('{}/{}_h{}'.format(results_dir, model, h))
            for m, p in hparams_store:
                y_hat, _, rmse, results_t = fl.ar_pls_expanding_window(h=h, p=p, yo_t=fl.yo_t,
                                                                       y_t=fl.y_t[:, h_idx][..., None],
                                                                       yo_v=fl.yo_v,
                                                                       y_v=fl.y_v[:, h_idx][..., None],
                                                                       rolling=rolling,
                                                                       save_dir=data_store_save_dir,
                                                                       save_name=model,
                                                                       )

                rmse_store.append(rmse)
                aic_t_store.append(results_t.aic)
                bic_t_store.append(results_t.bic)
                y_hat_store.append(y_hat)
            df = pd.DataFrame(data=np.concatenate((np.array(hparams_store),
                                                   np.array(rmse_store)[..., None],
                                                   np.array(aic_t_store)[..., None],
                                                   np.array(bic_t_store)[..., None],
                                                   np.array(y_hat_store)), axis=1),
                              columns=['m', 'p', 'Val RMSE', 'AIC_t', 'BIC_t'] + fl.y_v[:, h_idx].flatten().tolist())


        elif type == 'AIC_BIC':
            for m, p in hparams_store:
                aic, bic = fl.aic_bic_for_ar(h=h, p=p, h_max=h_max, p_max=p_max, yo=fl.yo,
                                             y=fl.y[:, h_idx][..., None])
                rmse_store.append(-1)
                aic_t_store.append(aic)
                bic_t_store.append(bic)
            df = pd.DataFrame(data=np.concatenate((np.array(hparams_store),
                                                   np.array(rmse_store)[..., None],
                                                   np.array(aic_t_store)[..., None],
                                                   np.array(bic_t_store)[..., None]), axis=1),
                              columns=['m', 'p', 'Val RMSE', 'AIC_t', 'BIC_t'])
    elif model in ['CW{}'.format(i) for i in range(1, 10)] + ['CWd{}'.format(i) for i in range(1, 10)] + [
        'XGB{}'.format(i) for i in range(1, 10)]:
        z_type = int(model[-1])
        cw_model_class = kwargs['cw_model_class']
        cw_hparams = kwargs['cw_hparams']
        if type == 'PLS':
            for m, p in hparams_store:
                data_store_save_dir = create_results_directory('{}/{}_h{}'.format(results_dir, model, h))
                y_hat, _, rmse, _ = fl.multicore_pls_expanding_window(h=h, p=p, m=m, r=r,
                                                                      cw_model_class=cw_model_class,
                                                                      cw_hparams=cw_hparams,
                                                                      x_t=fl.x_t,
                                                                      x_v=fl.x_v,
                                                                      yo_t=fl.yo_t,
                                                                      y_t=fl.y_t[:, h_idx][..., None],
                                                                      yo_v=fl.yo_v,
                                                                      y_v=fl.y_v[:, h_idx][..., None],
                                                                      rolling=rolling,
                                                                      z_type=z_type,
                                                                      save_dir=data_store_save_dir,
                                                                      save_name=model)

                rmse_store.append(rmse)
                y_hat_store.append(y_hat)

            aic_t_store = [np.nan] * len(rmse_store)
            bic_t_store = [np.nan] * len(rmse_store)
            df = pd.DataFrame(data=np.concatenate((np.array(hparams_store),
                                                   np.array(rmse_store)[..., None],
                                                   np.array(aic_t_store)[..., None],
                                                   np.array(bic_t_store)[..., None],
                                                   np.array(y_hat_store)), axis=1),
                              columns=['m', 'p', 'Val RMSE', 'AIC_t', 'BIC_t'] + fl.y_v[:, h_idx].flatten().tolist())

        elif type == 'k_fold':
            z_type = int(model[-1])
            results_df = fl.xgb_hparam_opt(x=fl.x_t, yo=fl.yo_t, y=fl.y_t[:, [h_idx]], h=h, m_max=m_max, p_max=p_max,
                                           z_type=z_type,
                                           hparam_opt_params=kwargs['hparam_opt_params'],
                                           default_hparams=cw_hparams,
                                           results_dir=results_dir,
                                           model_name=model)
            try:
                wb = openpyxl.load_workbook(f'{results_dir}/{model}_hparam_opt.xlsx')
            except FileNotFoundError:
                wb = openpyxl.Workbook()
            wb.create_sheet(f'{model}_h{h}')
            ws = wb[wb.sheetnames[-1]]
            print_df_to_excel(df=pd.DataFrame({'r': r,
                                               'm': m_max,
                                               'p': p_max},
                                              index=[0]),
                              ws=ws)
            print_df_to_excel(df=results_df, ws=ws, start_row=3)
            wb.save(f'{results_dir}/{model}_hparam_opt.xlsx')
            df = results_df
        elif type == 'AIC_BIC':
            raise KeyError('CW model selected. AIC_BIC method not available for CW model.')
    else:
        raise KeyError('Factor model selected is not available.')

    return df


class Fl_master:
    def __init__(self, x, yo, time_stamp, features_names, labels_names, y=None, y_names=None, time_idx=None):
        self.x = x  # 2D ndarray. The _o stands for original dataset for features and labels.
        if y is None:
            self.y = yo  # 2D ndarray
            self.y_names = None
        else:
            self.y = y
            self.y_names = y_names
        self.yo = yo  # 2D ndarray of y original without cumulative changes
        self.time_stamp = [f'{x}:{y}' for x, y in
                           zip(pd.DatetimeIndex(time_stamp).year, pd.DatetimeIndex(time_stamp).day)]  # 1D ndarray
        self.nobs, self.N = np.shape(x)  # Scalar X scalar
        if time_idx is not None:
            self.time_idx = time_idx
        else:
            self.time_idx = np.arange(self.nobs)  # 1D ndarray
        self.features_names = features_names
        self.labels_names = labels_names

        pass

    '''
    NOT USED
    def y_h_step_transformation(self, h_steps):
        y_store = []

        yo = self.yo.flatten()
        for h in h_steps:
            if type == 5:
                y_transformed = 1200 / h * np.log(yo[h:] / yo[:-h])
                y_transformed = [np.nan] * h + y_transformed.tolist()
                y_store.append(y_transformed)
            elif type == 6:
                y_transformed = np.array(yo)[2:] - 2 * np.array(yo)[1:-1] + np.array(yo)[:-2]
                y_transformed = [np.nan, np.nan] + y_transformed.tolist()
                y_store.append(y_transformed)
            else:
                raise KeyError('Label type is not 5 or 6')

        self.yo = self.yo * 1200
    '''

    def pca_factor_estimation(self, x, r, x_transformed_already=False):
        if not x_transformed_already:
            x_scaler = StandardScaler()
            x_scaler.fit(x)
            x = x_scaler.transform(x)

        w, v = eigh(x.T @ x)
        loadings = np.fliplr(v[:, -r:])
        loadings = loadings * math.sqrt(self.N)
        factors = x @ loadings / self.N
        loadings_T = loadings.T

        return factors, loadings_T

    def iterated_em(self, features, labels, pca_p, max_iter, tol, excel_dir):
        # all_x = np.concatenate((features, labels), axis=1)
        all_x = features
        nan_store = np.where(np.isnan(all_x))
        col_mean = np.nanmean(all_x, axis=0)
        all_x[nan_store] = np.take(col_mean, nan_store[1])

        n_nan = len(nan_store[0])
        diff = 1000
        iter = 0
        while iter < max_iter:
            if diff < tol:
                print('Tolerance met with maximum percentage error = {}%\n Number of iterations taken = {}'.format(
                    diff, iter))
                break
            all_x_scaler = StandardScaler()
            all_x_scaler.fit(all_x)
            all_x_norm = all_x_scaler.transform(all_x)

            factors, loadings_T = self.pca_factor_estimation(x=all_x_norm, r=pca_p, x_transformed_already=True)

            all_x_norm_1 = factors @ loadings_T
            all_x_1 = all_x_scaler.inverse_transform(all_x_norm_1)

            try:
                # diff = max((all_x_1[nan_store] - all_x[nan_store]) / all_x[nan_store] * 100)  # Tolerance is in percentage error
                diff = np.max((factors - factors_old) / factors_old * 100)
                factors_old = factors
            except:
                diff = 1000
                factors_old = factors
            all_x[nan_store] = all_x_1[nan_store]
            iter += 1
        else:
            raise ValueError(
                'During iterated EM method, maximum iteration of {} without meeting tolerance requirement.'
                ' Current maximum percentage error = {}%'.format(iter, diff))

        wb = openpyxl.load_workbook(excel_dir)
        wb.create_sheet('iter_em results')
        sheet_name = wb.sheetnames[-1]
        ws = wb[sheet_name]

        df = pd.DataFrame(data=np.concatenate((self.time_stamp[..., None], all_x), axis=1),
                          columns=self.features_names)

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        wb.save(excel_dir)
        return all_x

    def percentage_split(self, percentage):
        idx_split = round(self.nobs * (1 - percentage))
        return (self.x[:idx_split, :], self.x[idx_split:, :]), \
               (self.yo[:idx_split, :], self.yo[idx_split:, :]), \
               (self.y[:idx_split, :], self.y[idx_split:, :]), \
               (self.time_stamp[:idx_split], self.time_stamp[idx_split:]), \
               (self.time_idx[:idx_split], self.time_idx[idx_split:]), \
               (idx_split, self.nobs - idx_split)

    def date_split(self, date, date_start=None):
        '''
        Split data into two parts, first part is start date till date (inclusive), second part is everything after.
        :param date: date to split set 1 and set 2
        :param date_start: date to start set 1 from (exclusive)
        :return:
        '''
        idx_split = self.time_stamp.index(date) + 1
        if date_start:
            idx_start = self.time_stamp.index(date_start) + 1
            return (self.x[idx_start:idx_split, :], self.x[idx_split:, :]), \
                   (self.yo[idx_start:idx_split, :], self.yo[idx_split:, :]), \
                   (self.y[idx_start:idx_split, :], self.y[idx_split:, :]), \
                   (self.time_stamp[idx_start:idx_split], self.time_stamp[idx_split:]), \
                   (self.time_idx[idx_start:idx_split], self.time_idx[idx_split:]), \
                   (idx_split - idx_start - 1, self.nobs - idx_split)
        else:
            return (self.x[:idx_split, :], self.x[idx_split:, :]), \
                   (self.yo[:idx_split, :], self.yo[idx_split:, :]), \
                   (self.y[:idx_split, :], self.y[idx_split:, :]), \
                   (self.time_stamp[:idx_split], self.time_stamp[idx_split:]), \
                   (self.time_idx[:idx_split], self.time_idx[idx_split:]), \
                   (idx_split, self.nobs - idx_split)


class Fl_pca(Fl_master):
    def __init__(self, val_split, y, **kwargs):
        if val_split:
            super(Fl_pca, self).__init__(**kwargs)
            self.val_split = val_split
            self.y = y
            (self.x_t, self.x_v), (self.yo_t, self.yo_v), (self.y_t, self.y_v), (self.ts_t, self.ts_v), (
                self.tidx_t, self.tidx_v), (self.nobs_t, self.nobs_v) = self.percentage_split(val_split)

    def pca_factor_estimation(self, x, r, x_transformed_already=False):
        if not x_transformed_already:
            x_scaler = StandardScaler()
            x_scaler.fit(x)
            x = x_scaler.transform(x)
        '''
        pca = PCA(n_components=r)

        pca.fit_transform(x.T)
        factors = pca.components_.T * math.sqrt(self.nobs)
        loadings_T = (factors.T @ x / self.nobs)
        '''
        '''
        pca.fit_transform(x)
        loadings = pca.components_.T * math.sqrt(self.N)
        factors = x @ loadings / self.N
        '''
        T, N = x.shape
        w, v = eigh(x.T @ x)
        loadings = np.fliplr(v[:, -r:])
        loadings = loadings * math.sqrt(N)
        factors = x @ loadings / N
        loadings_T = loadings.T

        '''
        w, v = eigh(x @ x.T)
        factors = np.fliplr(v[:, -r:]) * math.sqrt(self.nobs)
        loadings_T = factors.T @ x / self.nobs
        '''

        return factors, loadings_T

    def umap_factor_estimation(self, x, r, x_transformed_already=False):
        if not x_transformed_already:
            x_scaler = StandardScaler()
            x_scaler.fit(x)
            x = x_scaler.transform(x)
        reducer = umap.UMAP(a=1.576943460405378, angular_rp_forest=False,
                            b=0.8950608781227859, init='spectral',
                            local_connectivity=1.0, metric='euclidean', metric_kwds={},
                            min_dist=0.1, n_components=r, n_epochs=None, n_neighbors=15,
                            negative_sample_rate=5, random_state=42, set_op_mix_ratio=1.0,
                            spread=1.0, target_metric='l2', target_metric_kwds={},
                            transform_queue_size=4.0, transform_seed=42, verbose=False)
        reducer.fit(x)
        factors = reducer.transform(x)

        return factors, None

    def ic_value(self, x, factors, loadings_T, x_transformed_already=False):
        if not x_transformed_already:
            x_scaler = StandardScaler()
            x_scaler.fit(x)
            x = x_scaler.transform(x)
        x_hat = factors @ loadings_T
        v = 1 / (self.nobs * self.N) * np.sum((x - x_hat) ** 2)
        k = np.shape(factors)[1]

        # Using g2 penalty.
        c_nt = min(self.nobs, self.N)
        return math.log(v) + k * ((self.nobs + self.N) / (self.nobs * self.N)) * math.log(c_nt)

    def pca_umap_prepare_data_matrix(self, factors, yo, y, h, m, p):
        '''

        :param factors: 2D ndarray of factors. observation x dimension = N x r
        :param y: 2D ndarray of y observation values. N x 1
        :param l: 2D ndarray of target y labels. Either same as y or cumulative difference vector of (N-h) x 1
        :param h: h steps ahead
        :param m: number of factor lags
        :param p: number of AR lags on Y
        :return: [ff, fy, l] = [factors data matrix, y data matrix, labels for target y h step ahead]
        '''
        a = max(m, p)
        T = np.shape(factors)[0]

        y = y[-T + h + a - 1:, :]
        ff = factors[a - 1:T - h, :]
        fy = yo[a - 1:T - h, :]


        if m >= 2:
            for idx in range(2, m + 1):
                ff = np.concatenate((ff, factors[a - idx + 1 - 1:T - h - idx + 1, :]), axis=1)

        if p >= 2:
            for idx in range(2, p + 1):
                fy = np.concatenate((fy, yo[a - idx + 1 - 1:T - h - idx + 1, :]), axis=1)

        return ff, fy, y

    '''
    NOT IN USE
    def prepare_data_vector(self, f, yo, m, p):
        v = f[-1, :][None, ...]

        if m >= 2:
            for idx in range(2, m + 1):
                v = np.concatenate((v, f[-idx, :][None, ...]), axis=1)

        if p >= 1:
            for idx in range(1, p + 1):
                v = np.concatenate((v, yo[-idx, :][None, ...]), axis=1)

        return v

    def prepare_validation_data_h_step_ahead(self, x_v, yo_v, y_v, h):
        if h == 1:
            return x_v, yo_v, y_v
        else:
            return x_v[:-(h-1), :], yo_v[:-(h-1), :], y_v[h-1:]
    '''

    def pls_expanding_window(self, h, m, p, r, factor_model, x_t, yo_t, y_t, x_v, yo_v, y_v,
                             rolling=False, save_dir=None, save_name=None):
        y_hat_store = []
        e_hat_store = []

        # x_v, yo_v, y_v = self.prepare_validation_data_h_step_ahead(x_v, yo_v, y_v, h)
        n_val = np.shape(x_v)[0]

        # Training on train set first
        f_t, _ = factor_model(x=x_t, r=r)
        f_LM_t, yo_LM_t, y_RO_t = self.pca_umap_prepare_data_matrix(f_t, yo_t, y_t, h, m, p)
        ols_model = sm.OLS(endog=y_RO_t, exog=sm.add_constant(np.concatenate((f_LM_t, yo_LM_t), axis=1)))

        ols_model = ols_model.fit()
        results_t = copy.deepcopy(ols_model)

        for idx, (x_1, yo_1, y_1) in enumerate(zip(x_v.tolist(), yo_v.tolist(), y_v.tolist())):
            x_t = np.concatenate((x_t, np.array(x_1)[None, ...]), axis=0)
            yo_t = np.concatenate((yo_t, np.array(yo_1)[None, ...]), axis=0)
            y_t = np.concatenate((y_t, np.array(y_1)[None, ...]), axis=0)
            f_t = np.concatenate((f_t, np.zeros((1, f_t.shape[1]))), axis=0)

            # f_t, _ = factor_model(x=x_t, r=r) SHOULD NOT RE-ESTIMATE FACTORS BEFORE FORECASTING
            f_LM_t, yo_LM_t, y_RO_t= self.pca_umap_prepare_data_matrix(f_t, yo_t, y_t, h, m, p)
            exog = np.concatenate((f_LM_t, yo_LM_t), axis=1)[-1, :][None, ...]

            y_1_hat = ols_model.predict(exog=np.concatenate((np.ones((1, 1)), exog), axis=1))
            e_1_hat = y_1 - y_1_hat

            y_hat_store.append(y_1_hat.item())
            e_hat_store.append(e_1_hat.item())

            if save_dir:
                try:
                    data_store.append({'y_hat': y_1_hat.item(),
                                       'e_hat': e_1_hat.item(),
                                       'aic': ols_model.aic,
                                       'bic': ols_model.bic})
                except:
                    data_store = [{'y_hat': y_1_hat.item(),
                                   'e_hat': e_1_hat.item(),
                                   'aic': ols_model.aic,
                                   'bic': ols_model.bic}]

            if idx + 1 == n_val:
                break  # since last iteration, no need to waste time re-estimating model

            if rolling:
                # Drop the first observation in the matrix for rolling window
                x_t = x_t[1:, :]
                yo_t = yo_t[1:, :]
                y_t = y_t[1:, :]
                f_t, _ = factor_model(x=x_t, r=r)
                f_LM_t, yo_LM_t, y_RO_t = self.pca_umap_prepare_data_matrix(f_t, yo_t, y_t, h, m, p)

            ols_model = sm.OLS(endog=y_RO_t, exog=sm.add_constant(np.concatenate((f_LM_t, yo_LM_t), axis=1)))
            ols_model = ols_model.fit()

        if save_dir:
            with open('{}/{}_m{}_p{}_h{}.pkl'.format(save_dir, save_name, m, p, h), "wb") as file:
                pickle.dump(data_store, file)

        return y_hat_store, e_hat_store, math.sqrt(np.mean(np.array(e_hat_store) ** 2)), results_t

    def aic_bic_for_pca_umap(self, h, m, p, r, h_max, m_max, p_max, factor_model, x, yo, y):
        f, _ = factor_model(x=x, r=r)
        f_LM, yo_LM, y = self.pca_umap_prepare_data_matrix(f, yo, y, h, m, p)
        T = np.shape(f)[0]
        a_max = max(m_max, p_max)
        if h < h_max or m < m_max or p < p_max:
            f_LM = f_LM[-T + h_max + a_max - 1:, :]
            yo_LM = yo_LM[-T + h_max + a_max - 1:, :]
            y = y[-T + h_max + a_max - 1:, :]
            pass

        ols_model = sm.OLS(endog=y, exog=sm.add_constant(np.concatenate((f_LM, yo_LM), axis=1)))
        ols_model = ols_model.fit()
        return ols_model.aic, ols_model.bic

    def pca_k_selection(self, lower_k=5, upper_k=100):
        ic_store = []
        x = self.x
        x_scaler = StandardScaler()
        x_scaler.fit(x)
        x = x_scaler.transform(x)

        # Training phase. First get best k value for factor estimation using IC criteria
        k_store = list(range(lower_k, upper_k + 1))
        for k in k_store:
            factors, loadings_T = self.pca_factor_estimation(x=x, r=k, x_transformed_already=True)
            ic_store.append(self.ic_value(x=x, factors=factors, loadings_T=loadings_T, x_transformed_already=True))

        r = k_store[np.argmin(ic_store)]

        print('Training results: Optimal factor dimension r = {}'.format(r))

        return r, ic_store

    def pca_hparam_opt(self, x, yo, y, h, m_max, p_max, r):
        def calculate_aic(x, yo, y, r, m, p):
            f_t, _ = self.pca_factor_estimation(x=x, r=r)
            f_LM_t, yo_LM_t, y_RO_t = self.pca_umap_prepare_data_matrix(f_t, yo, y, h, m, p)
            ols_model = sm.OLS(endog=y_RO_t, exog=sm.add_constant(np.concatenate((f_LM_t, yo_LM_t), axis=1)))
            ols_model = ols_model.fit()
            return ols_model.aic

        m_store = list(range(1, m_max + 1))
        p_store = list(range(1, p_max + 1))
        hparams_store = list(itertools.product(m_store, p_store))
        aic_store = []
        for (m,p) in hparams_store:
            aic_store.append(calculate_aic(x,yo,y,r,m,p))

        df = pd.DataFrame(data=np.concatenate((np.array(hparams_store), np.array(aic_store)[...,None]), axis=1),
                          columns=['m', 'p', 'AIC']).sort_values('AIC')

        return df


class Fl_ar(Fl_master):
    def __init__(self, val_split, y, **kwargs):
        if val_split:
            super(Fl_ar, self).__init__(**kwargs)
            self.val_split = val_split
            self.y = y
            (self.x_t, self.x_v), (self.yo_t, self.yo_v), (self.y_t, self.y_v), (self.ts_t, self.ts_v), (
                self.tidx_t, self.tidx_v), (self.nobs_t, self.nobs_v) = self.percentage_split(val_split)

    def ar_prepare_data_matrix(self, yo, y, h, p):
        '''

        :param factors: 2D ndarray of factors. observation x dimension = N x r
        :param y: 2D ndarray of y observation values. N x 1
        :param l: 2D ndarray of target y labels. Either same as y or cumulative difference vector of (N-h) x 1
        :param h: h steps ahead
        :param m: number of factor lags
        :param p: number of AR lags on Y
        :return: [ff, fy, l] = [factors data matrix, y data matrix, labels for target y h step ahead]
        '''
        a = p
        T = np.shape(yo)[0]

        y = y[-T + h + a - 1:, :]
        fy = yo[a - 1:T - h, :]

        if p >= 2:
            for idx in range(2, p + 1):
                fy = np.concatenate((fy, yo[a - idx + 1 - 1:T - h - idx + 1, :]), axis=1)

        return fy, y

    '''
    NOT IN USE
    def ar_prepare_data_vector(self, yo, p):
        v = yo[-1, :][None, ...]

        if p >= 2:
            for idx in range(2, p + 1):
                v = np.concatenate((v, yo[-idx, :][None, ...]), axis=1)

        return v

    def ar_prepare_validation_data_h_step_ahead(self, yo_v, y_v, h):
        if h == 1:
            return yo_v, y_v
        else:
            return yo_v[:-(h-1), :], y_v[h-1:]
    '''

    def ar_pls_expanding_window(self, h, p,  yo_t, y_t, yo_v, y_v, rolling=False,
                                save_dir=None, save_name=None):
        y_hat_store = []
        e_hat_store = []

        # yo_v, y_v = self.ar_prepare_validation_data_h_step_ahead(yo_v, y_v, h)
        n_val = np.shape(yo_v)[0]

        # Training on train set first.
        # yo = information set for original y. As new information comes in, append to this matrix.
        # y = information set for transformed y to be forecasted.
        # yo_LM_t = original y lag matrix for training dataset
        # y_RO_t = y regression output for training dataset
        yo_LM_t, y_RO_t = self.ar_prepare_data_matrix(yo_t, y_t, h, p)
        ols_model = sm.OLS(endog=y_RO_t, exog=sm.add_constant(yo_LM_t))

        ols_model = ols_model.fit()
        results_t = copy.deepcopy(ols_model)

        for idx, (yo_1, y_1) in enumerate(zip(yo_v.tolist(), y_v.tolist())):
            yo_t = np.concatenate((yo_t, np.array(yo_1)[None, ...]), axis=0)
            y_t = np.concatenate((y_t, np.array(y_1)[None, ...]), axis=0)

            yo_LM_t, y_RO_t = self.ar_prepare_data_matrix(yo_t, y_t, h, p)
            exog = yo_LM_t[-1, :][None, ...]

            y_1_hat = ols_model.predict(exog=np.concatenate((np.ones((1, 1)), exog), axis=1))
            e_1_hat = y_1 - y_1_hat

            y_hat_store.append(y_1_hat.item())
            e_hat_store.append(e_1_hat.item())

            if save_dir:
                try:
                    data_store.append({'y_hat': y_1_hat.item(),
                                       'e_hat': e_1_hat.item(),
                                       'aic': ols_model.aic,
                                       'bic': ols_model.bic})
                except:
                    data_store = [{'y_hat': y_1_hat.item(),
                                   'e_hat': e_1_hat.item(),
                                   'aic': ols_model.aic,
                                   'bic': ols_model.bic}]

            if idx + 1 == n_val:
                break  # since last iteration, no need to waste time re-estimating model

            if rolling:
                yo_t = yo_t[1:, :]
                y_t = y_t[1:, :]
                yo_LM_t = yo_LM_t[1:, :]
                y_RO_t = y_RO_t[1:, :]

            # yo_LM_t, y_t, y_idx_t = self.ar_prepare_data_matrix(yo_t, y_t, h, p)
            ols_model = sm.OLS(endog=y_RO_t, exog=sm.add_constant(yo_LM_t))
            ols_model = ols_model.fit()

        if save_dir:
            with open('{}/{}_p{}_h{}.pkl'.format(save_dir, save_name, p, h), "wb") as file:
                pickle.dump(data_store, file)

        return y_hat_store, e_hat_store, math.sqrt(np.mean(np.array(e_hat_store) ** 2)), results_t

    def ar_hparam_opt(self, yo, y, h, p_max):
        hparams_store = list(range(1, p_max + 1))
        aic_store = []
        for p in hparams_store:
            aic_store.append(self.aic_bic_for_ar(h=h, p=p, h_max=h, p_max=p_max, yo=yo,y=y)[0])

        df = pd.DataFrame(data=np.concatenate((np.array(hparams_store)[...,None], np.array(aic_store)[...,None]), axis=1),
                          columns=['p', 'AIC']).sort_values('AIC')

        return df

    def aic_bic_for_ar(self, h, p, h_max, p_max, yo, y):
        yo_LM, y = self.ar_prepare_data_matrix(yo, y, h, p)
        T = np.shape(yo_LM)[0]
        # Make sure that equal number of observations for all the samples
        a_max = p_max-p
        if h < h_max or p < p_max:
            yo_LM = yo_LM[-T + h_max + a_max - 1:, :]
            y = y[-T + h_max + a_max - 1:, :]
            pass

        ols_model = sm.OLS(endog=y, exog=sm.add_constant(yo_LM))
        ols_model = ols_model.fit()
        return ols_model.aic, ols_model.bic


class Fl_cw(Fl_master):
    def __init__(self, val_split, y, **kwargs):
        if val_split:
            super().__init__(**kwargs)
            self.val_split = val_split
            self.y = y
            (self.x_t, self.x_v), (self.yo_t, self.yo_v), (self.y_t, self.y_v), (self.ts_t, self.ts_v), (
                self.tidx_t, self.tidx_v), (self.nobs_t, self.nobs_v) = self.percentage_split(val_split)
        else:
            self.features_names = kwargs['features_names']

    def prepare_data_matrix(self, x, yo, y, h, m, p, z_type, feature_names=None):
        '''

        :param x: 2D ndarray of x. observation x dimension = N x r
        :param y: 2D ndarray of y observation values. N x 1
        :param l: 2D ndarray of target y labels. Either same as y or cumulative difference vector of (N-h) x 1
        :param h: h steps ahead
        :param m: number of factor lags
        :param p: number of AR lags on Y
        :return: [ff, fy, l] = [factors data matrix, y data matrix, labels for target y h step ahead]
        '''
        if z_type == 1:
            z = x
        elif z_type == 2:
            z = np.concatenate((x, x ** 2), axis=1)
        elif z_type == 3:
            z, _ = self.pca_factor_estimation(x=x, r=8)
        elif z_type == 4:
            z, _ = self.pca_factor_estimation(x=x, r=8)
            z = np.concatenate((z, z ** 2), axis=1)
        elif z_type == 5:
            s = pd.DataFrame(x)
            z = np.array(pd.concat([s, s.shift(), s.shift(2), s.shift(3), s.shift(4)], axis=1))[4:, :]
            yo = yo[4:, :]
            y = y[4:, :]
            z, _ = self.pca_factor_estimation(x=np.concatenate((z, z ** 2), axis=1), r=8)
        elif z_type == 6:
            s = pd.DataFrame(x)
            z = np.array(pd.concat([s, s.shift(), s.shift(2), s.shift(3), s.shift(4)], axis=1))[4:, :]
            yo = yo[4:, :]
            y = y[4:, :]
            z, _ = self.pca_factor_estimation(x=np.concatenate((z, z ** 2), axis=1), r=8)
            z = np.concatenate((z, z ** 2), axis=1)
        elif z_type == 7:
            z, _ = self.pca_factor_estimation(x=np.concatenate((x, x ** 2), axis=1), r=8)
        elif z_type == 8:
            z, _ = self.pca_factor_estimation(x=np.concatenate((x, x ** 2), axis=1), r=8)
            z = np.concatenate((z, z ** 2), axis=1)
        elif z_type == 9:
            try:
                cut = np.where(np.isnan(y).squeeze())[0][-1] + 1
                z = x[cut:, :]
                yo = y[cut:, :]
                y = y[cut:, :]
            except IndexError:
                z = x
                yo = y

        a = max(m, p)
        T = np.shape(z)[0]

        y = y[-T + h + a - 1:, :]
        ff = z[a - 1:T - h, :]
        fy = yo[a - 1:T - h, :]

        if m >= 2:
            for idx in range(2, m + 1):
                ff = np.concatenate((ff, z[a - idx + 1 - 1:T - h - idx + 1, :]), axis=1)

        if p >= 2:
            for idx in range(2, p + 1):
                fy = np.concatenate((fy, yo[a - idx + 1 - 1:T - h - idx + 1, :]), axis=1)

        if feature_names is not None:
            names = ['constant'] + [f'{feature}_L{idx}' for idx in range(m) for feature in feature_names] + [f'y_L{idx}'
                                                                                                             for idx in
                                                                                                             range(p)]
            return sm.add_constant(np.concatenate((ff, fy), axis=1)), y, names
        else:
            return sm.add_constant(np.concatenate((ff, fy), axis=1)), y

    def pls_expanding_window(self, h, m, p, cw_model_class, cw_hparams, x_t, yo_t, y_t, x_v, yo_v, y_v, z_type,
                             rolling=False,
                             save_dir=None, save_name=None):
        y_hat_store = []
        e_hat_store = []

        n_val = np.shape(x_v)[0]

        # Training on train set first
        z_matrix, y_vec = self.prepare_data_matrix(x_t, yo_t, y_t, h, m, p, z_type)
        cw_model = cw_model_class(z_matrix=z_matrix, y_vec=y_vec, hparams=cw_hparams)
        cw_model.fit()

        for idx, (x_1, yo_1, y_1) in enumerate(zip(x_v.tolist(), yo_v.tolist(), y_v.tolist())):
            x_t = np.concatenate((x_t, np.array(x_1)[None, ...]), axis=0)
            yo_t = np.concatenate((yo_t, np.array(yo_1)[None, ...]), axis=0)
            y_t = np.concatenate((y_t, np.array(y_1)[None, ...]), axis=0)

            z_matrix, y_vec = self.prepare_data_matrix(x_t, yo_t, y_t, h, m, p, z_type)
            exog = z_matrix[-1:, :]

            y_1_hat = cw_model.predict(exog=exog)
            e_1_hat = y_1[0] - y_1_hat.item()

            y_hat_store.append(y_1_hat.item())
            e_hat_store.append(e_1_hat)

            if save_dir:
                end = time.time()
                if (idx) % 5 == 0:
                    try:
                        print('Time taken for 5 steps CW PLS is {}'.format(end - start))
                        with open('{}/{}_h{}_{}.pkl'.format(save_dir, save_name, h, idx), "wb") as file:
                            pickle.dump(data_store, file)
                    except:
                        pass
                    data_store = []
                data_store.append(cw_model.return_data_dict())
                start = time.time()

            if idx + 1 == n_val:
                break  # since last iteration, no need to waste time re-estimating model

            if rolling:
                # Drop the first observation in the matrix for rolling window
                x_t = x_t[1:, :]
                yo_t = yo_t[1:, :]
                y_t = y_t[1:, :]
                z_matrix, y_vec = self.prepare_data_matrix(x_t, yo_t, y_t, h, m, p, z_type)

            cw_model = cw_model_class(z_matrix=z_matrix, y_vec=y_vec, hparams=cw_hparams)
            cw_model.fit()

        if idx % 5 != 0:
            with open('{}/{}_h{}_{}.pkl'.format(save_dir, save_name, h, idx), "wb") as file:
                pickle.dump(data_store, file)

        return y_hat_store, e_hat_store, math.sqrt(np.mean(np.array(e_hat_store) ** 2))

    def one_step_eval(self, arg, extra, idx):
        x, yo, y = arg
        cw_model_class, cw_hparams, h, m, p, r, z_type, save_dir = extra
        z_matrix, y_vec = self.prepare_data_matrix(x[:-1, :], yo[:-1, :], y[:-1, :], h, m, p, z_type)
        cw_model = cw_model_class(z_matrix=z_matrix, y_vec=y_vec, hparams=cw_hparams, r=r)
        if save_dir:
            plot_name = f'{save_dir}/{idx}'
        else:
            plot_name = None
        cw_model.fit(plot_name=plot_name)

        z_matrix, y_vec = self.prepare_data_matrix(x, yo, y, h, m, p, z_type)
        exog = z_matrix[-1:, :]

        y_1_hat = cw_model.predict(exog=exog).item()
        e_1_hat = y[-1].item() - y_1_hat

        if save_dir:
            t1 = time.perf_counter()
            e_1_hat_store = np.cumsum(cw_model.bhat_new_store.toarray(), axis=0) @ exog.squeeze() - y[-1].item()
            plt.plot(e_1_hat_store)
            plt.axvline(cw_model.m_star, linestyle='--')
            plt.savefig(f'{save_dir}/test_{idx}.png')
            plt.close()
            t2 = time.perf_counter()
            print(t2 - t1)

        return y_1_hat, e_1_hat, cw_model.return_data_dict()

    def multicore_pls_expanding_window(self, h, m, p, r, cw_model_class, cw_hparams, x_t, yo_t, y_t, x_v, yo_v, y_v,
                                       z_type,
                                       save_dir, save_name,
                                       rolling,
                                       ):
        if z_type < 3:
            # Only dataset 3,4,5,6,7,8 has factors
            r = None

        if rolling:
            args = [
                [np.concatenate((t[idx:, :], v[:idx + 1, :]), axis=0) for t, v in [[x_t, x_v], [yo_t, yo_v], [y_t, y_v]]
                 ] for idx in range(x_v.shape[0])]
        else:
            args = [[np.concatenate((t, v[:idx + 1, :]), axis=0) for t, v in [[x_t, x_v], [yo_t, yo_v], [y_t, y_v]]
                     ] for idx in range(x_v.shape[0])]

        with concurrent.futures.ProcessPoolExecutor(max_workers=4) as executer:
            results = executer.map(self.one_step_eval, args,
                                   itertools.repeat((cw_model_class, cw_hparams, h, m, p, r, z_type,
                                                     None)),
                                   itertools.count())

        y_hat_store, e_hat_store, cw_data_store = zip(*list(results))
        hparams = cw_hparams.copy()
        hparams.update([('h', h),
                        ('m', m),
                        ('p', p),
                        ('r', r),
                        ('z_type', z_type)])
        cw_data_store[0]['hparams'] = hparams
        if save_dir:
            with open('{}/{}_h{}_all.pkl'.format(save_dir, save_name, h), "wb") as file:
                pickle.dump(cw_data_store, file)

        return y_hat_store, e_hat_store, math.sqrt(np.mean(np.array(e_hat_store) ** 2)), cw_data_store


class Fl_xgb(Fl_cw):
    def pls_expanding_window(self, h, m, p, r, cw_model_class, cw_hparams, x_t, yo_t, y_t, x_v, yo_v, y_v,
                             z_type,
                             save_dir, save_name,
                             rolling=False):
        y_hat_store = []
        e_hat_store = []
        data_store = []
        n_val = np.shape(x_v)[0]
        x = x_t
        yo = yo_t
        y = y_t

        for idx, (x_1, yo_1, y_1) in enumerate(zip(x_v.tolist(), yo_v.tolist(), y_v.tolist())):
            x = np.concatenate((x, np.array(x_1)[None, ...]), axis=0)
            yo = np.concatenate((yo, np.array(yo_1)[None, ...]), axis=0)
            y = np.concatenate((y, np.array(y_1)[None, ...]), axis=0)

            z_matrix, y_vec, z_names = self.prepare_data_matrix(x[:-1, :], yo[:-1, :], y[:-1, :], h, m, p, z_type,
                                                                feature_names=self.features_names)
            cw_model = cw_model_class(z_matrix=z_matrix, y_vec=y_vec, hparams=cw_hparams, r=r)
            z_matrix, y_vec = self.prepare_data_matrix(x, yo, y, h, m, p, z_type)
            exog = z_matrix[-1:, :]
            cw_model.fit(deval=xgb.DMatrix(data=exog, label=y[[-1]]), ehat_eval=cw_hparams['ehat_eval'],
                         plot_name=None, feature_names=z_names)
            y_1_hat = cw_model.predict(exog=exog, best_ntree_limit=cw_model.model.best_ntree_limit).item()
            e_1_hat = y[-1].item() - y_1_hat

            y_hat_store.append(y_1_hat)
            e_hat_store.append(e_1_hat)

            data_store.append(cw_model.return_data_dict())

            if idx + 1 == n_val:
                break  # since last iteration, no need to waste time re-estimating model

            if rolling:
                # Drop the first observation in the matrix for rolling window
                x = x[1:, :]
                yo = yo[1:, :]
                y = y[1:, :]

        data_store[0]['feature_names'] = z_names
        if save_dir:
            with open('{}/{}_h{}.pkl'.format(save_dir, save_name, h), "wb") as file:
                pickle.dump(data_store, file)

        return y_hat_store, e_hat_store, math.sqrt(np.mean(np.array(e_hat_store) ** 2)), data_store

    def one_step_eval(self, arg, extra, idx):
        x, yo, y = arg
        cw_model_class, cw_hparams, h, m, p, r, z_type, save_dir = extra
        z_matrix, y_vec = self.prepare_data_matrix(x[:-1, :], yo[:-1, :], y[:-1, :], h, m, p, z_type)
        cw_model = cw_model_class(z_matrix=z_matrix, y_vec=y_vec, hparams=cw_hparams, r=r)
        if save_dir:
            plot_name = f'{save_dir}/{idx}'
        else:
            plot_name = None

        z_matrix, y_vec = self.prepare_data_matrix(x, yo, y, h, m, p, z_type)
        exog = z_matrix[-1:, :]

        cw_model.fit(deval=xgb.DMatrix(data=exog, label=y[[-1]]), ehat_eval=cw_hparams['ehat_eval'],
                     plot_name=plot_name)

        y_1_hat = cw_model.predict(exog=exog, best_ntree_limit=cw_model.model.best_ntree_limit).item()
        e_1_hat = y[-1].item() - y_1_hat

        if save_dir:
            t1 = time.perf_counter()
            e_1_hat_store = np.cumsum(cw_model.bhat_new_store.toarray(), axis=0) @ exog.squeeze() - y[-1].item()
            plt.plot(e_1_hat_store)
            plt.axvline(cw_model.m_star, linestyle='--')
            plt.savefig(f'{save_dir}/test_{idx}.png')
            plt.close()
            t2 = time.perf_counter()
            print(t2 - t1)

        return y_1_hat, e_1_hat, cw_model.return_data_dict()

    def val_rep_holdout(self, x, yo, y, h, z_type, n_blocks, hparams, **kwargs):
        '''
        rep_holdout evaluation method based on "Evaluating time series forecasting models".
        Splits into train and val set by selecting n_blocks points on evenly spaced intervals between 60-90%.
        For each train val pair, the xgb model is trained on train set and tested on the val set with early stopping.
        The average optimal round is the average optimal rounds of all runs.
        The average loss is calculated from all runs as well.
        :param z:
        :param y:
        :param n_blocks:
        :return:
        '''
        cut_points = np.linspace(start=0.75, stop=0.95, num=n_blocks)
        z, y = self.prepare_data_matrix(x, yo, y, h, hparams['m'], hparams['m'] * 2, z_type)
        T = z.shape[0]
        score = []
        n_rounds = []
        for idx, cut in enumerate(cut_points):
            cut = int(round(T * cut))
            z_train = z[:cut, :]
            z_test = z[cut:, :]
            y_train = y[:cut, :]
            y_test = y[cut:, :]
            cw_model = Xgboost(z_matrix=z_train, y_vec=y_train, hparams=hparams, r=None)
            cw_model.fit(deval=xgb.DMatrix(data=z_test, label=y_test), plot_name=None)  # f'./results/{idx}.png')
            y_hat = cw_model.predict(exog=z_test, best_ntree_limit=cw_model.model.best_ntree_limit)
            e_hat = y_test.squeeze() - y_hat
            score.append(np.sqrt(np.mean(e_hat ** 2)))
            n_rounds.append(cw_model.model.best_ntree_limit)
        return np.mean(score), int(round(np.mean(n_rounds)))

    def val_prequential(self, x, yo, y, h, z_type, cut_point, hparams, **kwargs):
        '''
        Prequential evaluation method based on "Evaluating time series forecasting models".
        Similar to expanding window out of sample forecast.
        Select one hparam then test it out on a single expanding window into the test set.
        :param z:
        :param y:
        :param n_blocks:
        :return:
        '''
        cut = int(round(x.shape[0] * cut_point))
        t0 = time.perf_counter()
        y_hat, _, rmse, cw_data_store = self.pls_expanding_window(h=h, p=hparams['m'] * 2, m=hparams['m'],
                                                                  r=8,
                                                                  cw_model_class=Xgboost,
                                                                  cw_hparams=hparams,
                                                                  x_t=x[:cut, :],
                                                                  x_v=x[cut:, :],
                                                                  yo_t=yo[:cut, :],
                                                                  y_t=y[:cut, :],
                                                                  yo_v=yo[cut:, :],
                                                                  y_v=y[cut:, :],
                                                                  rolling=False,
                                                                  z_type=z_type,
                                                                  save_dir=kwargs['save_dir'],
                                                                  save_name=kwargs['save_name'])
        rounds = int(round(np.mean([data['best_ntree_limit'] for data in cw_data_store])))
        t1 = time.perf_counter()
        print(f'Time taken for 1 hparam opt trial is {t1 - t0}')
        return rmse, rounds

    def xgb_hparam_opt(self, x, yo, y, h, m_max, p_max, z_type, hparam_opt_params, default_hparams, results_dir,
                       model_name):
        y_all = y.copy()
        # Space should include 1) max_depth, 2) colsample_bytree
        space = []
        for k, v in hparam_opt_params['variables'].items():
            if v['type'] == 'Real':
                space.append(Real(v['lower'], v['upper'], name=k))
            elif v['type'] == 'Integer':
                space.append(Integer(v['lower'], v['upper'], name=k))
            else:
                raise TypeError('hparam opt bounds variable type must be Real or Integer only.')
        n_rounds_store = []

        x_iters = []
        func_vals = []

        @use_named_args(space)
        def objective(**params):
            try:
                idx = x_iters.index(params)
                score, rounds = func_vals[idx]
                n_rounds_store.append(rounds)
                print(f'Re-evaluated {params}')
            except ValueError:
                x_iters.append(params)
                # Merge default hparams with optimizer trial hparams
                params = {**default_hparams, **params}
                if hparam_opt_params['val_mode'] == 'rfcv':
                    z, y = self.prepare_data_matrix(x, yo, y_all, h, params['m'], params['m']*2, z_type)
                    cv_results = xgb.cv(params=params, dtrain=xgb.DMatrix(data=z, label=y),
                                        nfold=hparam_opt_params['n_blocks'], num_boost_round=params['num_boost_round'],
                                        early_stopping_rounds=params['early_stopping_rounds'],
                                        metrics='rmse', as_pandas=True, seed=params['seed'])
                    rounds = cv_results.shape[0]
                    n_rounds_store.append(rounds)
                    score = cv_results['test-rmse-mean'].values[-1]
                    func_vals.append((score, rounds))
                elif hparam_opt_params['val_mode'] == 'rep_holdout':
                    score, rounds = self.val_rep_holdout(x, yo, y_all, h, z_type, n_blocks=hparam_opt_params['n_blocks'],
                                                         hparams=params)
                    n_rounds_store.append(rounds)
                    func_vals.append((score, rounds))
                elif hparam_opt_params['val_mode'] == 'prequential':
                    score, rounds = self.val_prequential(x, yo, y_all, h, z_type, cut_point=hparam_opt_params['cut_point'],
                                                         hparams=params,
                                                         save_dir=results_dir,
                                                         save_name=model_name)
                    n_rounds_store.append(rounds)
                    func_vals.append((score, rounds))
                else:
                    raise TypeError('hparam opt params val mode is invalid.')
            return score

        res_gp = gp_minimize(objective, space, random_state=default_hparams['seed'], acq_func='EI',
                             n_calls=hparam_opt_params['n_calls'],
                             n_random_starts=hparam_opt_params['n_random_starts'],
                             )
        df = pd.DataFrame(data=np.concatenate((res_gp.x_iters,
                                               np.array(n_rounds_store)[:, None],
                                               res_gp.func_vals[:, None]),
                                              axis=1),
                          columns=[s.name for s in res_gp.space] + ['m iters', 'val rmse']
                          ).sort_values('val rmse')
        return df


class Fl_cw_data_store(Fl_cw):
    def predict(self, b_hat_store, exog):
        return np.cumsum(b_hat_store.toarray(), axis=0) @ exog.squeeze()

    def pls_expanding_window(self, h, m, p, data_store, x_t, yo_t, y_t, x_v, yo_v, y_v, z_type, save_dir,
                             rolling=False):
        y_hat_store = []
        e_hat_store = []
        n_val = np.shape(x_v)[0]

        plot_idx = 20

        for idx, (x_1, yo_1, y_1) in enumerate(zip(x_v.tolist(), yo_v.tolist(), y_v.tolist())):
            x_t = np.concatenate((x_t, np.array(x_1)[None, ...]), axis=0)
            yo_t = np.concatenate((yo_t, np.array(yo_1)[None, ...]), axis=0)
            y_t = np.concatenate((y_t, np.array(y_1)[None, ...]), axis=0)

            z_matrix, y_vec = self.prepare_data_matrix(x_t, yo_t, y_t, h, m, p, z_type)
            exog = z_matrix[-1:, :]
            cum_bhat_store = np.cumsum(data_store[idx]['bhat_new_store'].toarray(), axis=0)
            y_1_hat = np.array(cum_bhat_store @ exog.squeeze())
            e_1_hat = y_1[0] - y_1_hat

            y_hat_store.append(y_1_hat)
            e_hat_store.append(e_1_hat)

            if idx % 20 == 0:
                # plot stuff
                train_error = np.mean((z_matrix @ cum_bhat_store.T - y_vec) ** 2, axis=0)
                plt.plot(train_error)
                plt.axvline(data_store[idx]['m_star'], linestyle='--', color='r')
                plt.ylim([min(train_error[:data_store[idx]['m_star'] + 2]),
                          max(train_error[:data_store[idx]['m_star'] + 2])])
                plt.savefig(f'{save_dir}/trainerror_{idx}.png')
                plt.close()

                plt.plot(e_1_hat)
                plt.axvline(data_store[idx]['m_star'], linestyle='--', color='r')
                plt.ylim([min(e_1_hat[:data_store[idx]['m_star'] + 2]), max(e_1_hat[:data_store[idx]['m_star'] + 2])])
                plt.savefig(f'{save_dir}/testerror_{idx}.png')
                plt.close()

            if idx + 1 == n_val:
                break  # since last iteration, no need to waste time re-estimating model

            if rolling:
                # Drop the first observation in the matrix for rolling window
                x_t = x_t[1:, :]
                yo_t = yo_t[1:, :]
                y_t = y_t[1:, :]

        return y_hat_store, e_hat_store, None
