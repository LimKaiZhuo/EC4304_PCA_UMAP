import pickle, os
import numpy as np
import pandas as pd
import openpyxl
from natsort import natsorted
import matplotlib.pyplot as plt
import seaborn as sns
import statsmodels.api as sm
import collections

from own_package.features_labels import read_excel_data, read_excel_dataloader, Fl_master, Fl_cw_data_store
from own_package.postprocess import read_excel_to_df
from own_package.others import print_df_to_excel


def prepare_grand_data_store(dir_store, model):
    data_store = []
    if isinstance(dir_store, str):
        dir_store = [dir_store]

    if model == 'AR' or model == 'PLS':
        # Each pickle datastore is one complete OOS forecasting but with different lags (m or p).
        for dir_name in dir_store:
            filenames = natsorted(os.listdir(dir_name))
            for filename in filenames:
                if filename.endswith(".pkl"):
                    with open('{}/{}'.format(dir_name, filename), 'rb') as handle:
                        data_store.append(pickle.load(handle))
    elif model == 'CW' or model == 'CWd':
        # each pickle datastore is 5 OOS forecast steps. Extend them so that it is one list.
        for dir_name in dir_store:
            filenames = natsorted(os.listdir(dir_name))
            for filename in filenames:
                if filename.endswith(".pkl"):
                    with open('{}/{}'.format(dir_name, filename), 'rb') as handle:
                        data_store.extend(pickle.load(handle))
    return data_store


def cw_analysis(**kwargs):
    excel_dir = kwargs['excel_dir']
    results_dir = kwargs['results_dir']
    rolling = kwargs['rolling']
    h_idx = kwargs['h_idx']
    h = kwargs['h']
    p = kwargs['p']
    m = kwargs['m']
    z_type = kwargs['z_type']
    skip_first_val = kwargs['skip_first_val']
    output = read_excel_dataloader(excel_dir=excel_dir)
    data_store = prepare_grand_data_store([results_dir], model='CW')

    if skip_first_val:
        # old implementation of CW class did not save the first model instance nor the last 4 instances.
        data_store.insert(0, data_store[0])
        skip_idx = 1
    else:
        skip_idx = 0

    fl_master = Fl_master(x=output[0], features_names=output[1],
                          yo=output[2], labels_names=output[3],
                          y=output[4], y_names=output[5],
                          time_stamp=output[6])
    (f_tv, f_tt), (yo_tv, yo_t), (y_tv, y_tt), \
    (ts_tv, ts_tt), (tidx_tv, tidx_tt), (nobs_tv, nobs_tt) = fl_master.percentage_split(0)

    fl = Fl_cw_data_store(val_split=0.2, x=f_tv, yo=yo_tv, y=y_tv,
                          time_stamp=ts_tv, time_idx=tidx_tv,
                          features_names=fl_master.features_names, labels_names=fl_master.labels_names,
                          y_names=fl_master.y_names)

    y_hat, e_hat, _ = fl.pls_expanding_window(h=h, p=p, m=m,
                                              data_store=data_store,
                                              x_t=fl.x_t,
                                              x_v=fl.x_v,
                                              yo_t=fl.yo_t,
                                              y_t=fl.y_t[:, h_idx][..., None],
                                              yo_v=fl.yo_v,
                                              y_v=fl.y_v[:, h_idx][..., None],
                                              z_type=z_type,
                                              save_dir=results_dir,
                                              rolling=rolling, )

    m_star_store = [x['m_star'] for x in data_store]
    y_hat_star = [y[i_star - skip_idx] for y, i_star in zip(y_hat, m_star_store)]
    e_hat_star = [e[i_star - skip_idx] for e, i_star in zip(e_hat, m_star_store)]

    e_df = pd.DataFrame(e_hat).iloc[:, :]
    e_df['month'] = e_df.index

    x = e_df.index
    e_df = pd.melt(e_df, id_vars='month', var_name='m', value_name='e_hat')
    sns.lineplot(x='month', y='e_hat', hue='m', data=e_df)
    plt.plot(x, e_hat_star)
    plt.savefig('{}/e_hat.png'.format(results_dir))
    plt.close()
    sns.lineplot(x='month', y='e_hat', hue='m', data=e_df.loc[e_df['m'] <= 70])
    plt.plot(x, e_hat_star)
    plt.savefig('{}/e_hat_m70.png'.format(results_dir))
    plt.close()
    plt.plot(m_star_store)
    plt.savefig('{}/m_star_store.png'.format(results_dir))
    plt.close()


class LocalLevel(sm.tsa.statespace.MLEModel):
    _start_params = [1., 1.]
    _param_names = ['var.level', 'var.irregular']

    def __init__(self, endog):
        super(LocalLevel, self).__init__(endog, k_states=1, initialization='diffuse')

        self['design', 0, 0] = 1
        self['transition', 0, 0] = 1
        self['selection', 0, 0] = 1

    def transform_params(self, unconstrained):
        return unconstrained ** 2

    def untransform_params(self, unconstrained):
        return unconstrained ** 0.5

    def update(self, params, **kwargs):
        params = super(LocalLevel, self).update(params, **kwargs)

        self['state_cov', 0, 0] = params[0]
        self['obs_cov', 0, 0] = params[1]


def single_xgb_analysis(results_dir, plot_name):
    data_store = prepare_grand_data_store([results_dir], model='CW')
    test_rmse_store = np.array([data['progress']['h_step_ahead']['rmse'] for data in data_store])
    best_idx_store = list(np.argmin(test_rmse_store, axis=1))
    predicted_idx = [30] * 5
    current_window = best_idx_store[:5]
    for y in best_idx_store[5:]:
        model = LocalLevel(current_window)
        res = model.fit(disp=False)
        predicted_idx.append(int(res.forecast()))
        current_window.append(y)
    plt.plot(best_idx_store, label='Orcale', marker='o', markersize=3)
    plt.plot(predicted_idx, label='Predicted', marker='o', markersize=3)
    plt.ylabel('m iteration')
    plt.xlabel('Sample')
    plt.legend()
    plt.savefig(f'{plot_name}_m.png')
    plt.close()
    T = len(test_rmse_store)
    m_max = len(test_rmse_store[0])
    orcale_ae = np.array([abs(single[idx]) for single, idx in zip(test_rmse_store, best_idx_store)])
    predicted_ae = np.array([abs(single[idx]) for single, idx in zip(test_rmse_store, predicted_idx)])
    m_max_ae = np.array([abs(single[-1]) for single, idx in zip(test_rmse_store, best_idx_store)])
    m0_ae = np.array([abs(single[0]) for single, idx in zip(test_rmse_store, best_idx_store)])
    plt.plot(predicted_ae-orcale_ae, label='Predicted')
    plt.plot(m_max_ae-orcale_ae, label=f'm = {m_max}')
    plt.ylabel('Abs Error compared to Orcale Abs Error')
    plt.xlabel('Sample')
    plt.legend()
    plt.savefig(f'{plot_name}_abserror.png')
    plt.close()

    fig = plt.figure()
    host = fig.add_subplot(111)

    par1 = host.twinx()
    par2 = host.twinx()

    #host.set_xlim(0, 2)
    #host.set_ylim(0, 2)
    #par1.set_ylim(0, 4)
    #par2.set_ylim(1, 65)

    host.set_xlabel("m Iterations")
    host.set_ylabel('Timestep 0')
    par1.set_ylabel(f"Timestep {int(T/2)}")
    par2.set_ylabel(f"Timestep {T}")

    color1 = plt.cm.viridis(0)
    color2 = plt.cm.viridis(0.5)
    color3 = plt.cm.viridis(.9)

    handles = []

    for t, subplot, color in zip([0, int(T/2), T-1],[host, par1, par2], [color1, color2, color3]):
        handles.extend(subplot.plot(test_rmse_store[t], label=f'Timestep {t} RMSE', color=color))
        subplot.scatter(best_idx_store[t], test_rmse_store[t,best_idx_store[t]], color='g', marker='o')
        subplot.scatter(predicted_idx[t], test_rmse_store[t, predicted_idx[t]], color='r', marker='x')

    # right, left, top, bottom
    par2.spines['right'].set_position(('outward', 60))
    # no x-ticks
    # par2.xaxis.set_ticks([])
    # Sometimes handy, same for xaxis
    # par2.yaxis.set_ticks_position('right')

    for handle, subplot in zip(handles, [host, par1, par2]):
        subplot.yaxis.label.set_color(handle.get_color())

    plt.savefig(f'{plot_name}_error_vs_m.png', bbox_inches='tight')
    plt.close()

    df = pd.DataFrame.from_dict({'Timestep':list(range(T)),
                                 'Orcale idx': best_idx_store,
                                 'Predicted idx': predicted_idx,
                                 'Orcale AE': orcale_ae,
                                 'Predicted AE': predicted_ae,
                                 'm_max AE': m_max_ae,
                                 'm0 AE': m0_ae})
    return np.mean([single[idx] ** 2 for single, idx in zip(test_rmse_store, predicted_idx)]) ** 0.5, df


def xgb_analysis(results_dir):
    h_steps = [f'h{h}' for h in [1, 3,6,12,24]]
    results_store = dict()
    for h in h_steps:
        results_store[h] = dict()
        for file in natsorted(os.listdir(results_dir)):
            if h == file.partition('_')[-1]:
                model_name = file.rpartition('/')[-1].partition('_')[0]
                rmse, df = single_xgb_analysis(
                    f'{results_dir}/{file}', plot_name=f'{results_dir}/{h}_{model_name}')
                results_store[h][model_name] = rmse
    df = pd.concat({k: pd.DataFrame(v, index=[0]).T.sort_values(0) for k, v in results_store.items()}, axis=0)
    df.to_excel(f'{results_dir}/best_xgb.xlsx')
    print('Done')


def combine_rmse_results(results_dir):
    df_store = []
    df_best_store = []
    for excel in os.listdir(results_dir):
        if 'testset_' in excel:
            print(f'Loading excel from {results_dir}/{excel}')
            df_store.append(read_excel_to_df(f'{results_dir}/{excel}'))
            df_best = []
            for df in df_store[-1]:
                df.insert(0, 'model', excel.split('_')[2])
                if '_AR_' in excel:
                    df_best.append(df.iloc[[3, df['Val RMSE'].argmin()], :])
                elif '_PCA_' in excel:
                    df_best.append(df.iloc[[df['Val RMSE'].argmin()], :])
                else:
                    df_best.append(df)
            df_best_store.append(df_best)
    # transpose nested list
    df_store = list(map(list, zip(*df_store)))
    combined_df_store = []
    for df_h in df_store:
        combined_df_store.append(pd.concat(df_h).sort_values(by='Val RMSE'))

    wb = openpyxl.Workbook()
    for h, df in zip([1, 3, 6, 12, 24], combined_df_store):
        wb.create_sheet(f'h_{h}')
        ws = wb[f'h_{h}']
        print_df_to_excel(df=df, ws=ws)

    wb.save(f'{results_dir}/summary.xlsx')

    # transpose nested list
    df_store = list(map(list, zip(*df_best_store)))
    combined_df_store = []
    for df_h in df_store:
        combined_df_store.append(pd.concat(df_h).sort_values(by='Val RMSE'))

    wb = openpyxl.Workbook()
    for h, df in zip([1, 3, 6, 12, 24], combined_df_store):
        wb.create_sheet(f'h_{h}')
        ws = wb[f'h_{h}']
        print_df_to_excel(df=df, ws=ws)

    wb.save(f'{results_dir}/best summary.xlsx')


def combine_best_summary_and_xgbs(best_summary_dir, xgbs_dir, results_dir):
    xgbs_df = pd.read_excel(xgbs_dir)
    bs_df_store = read_excel_to_df(best_summary_dir)
    combined_df_store = []
    columns = xgbs_df.columns
    h_rows = list(np.where(xgbs_df[columns[0]].notnull())[0])+[len(xgbs_df[columns[0]])]
    for h_start, h_end, bs_df in zip(h_rows[:-1], h_rows[1:], bs_df_store):
        count = h_end-h_start
        df = bs_df.reindex(bs_df.index.tolist() + list(range(90, 90+count))).reset_index(drop=True)
        model_names = [f'{x}s' for x in xgbs_df.iloc[h_start:h_end,1].values]
        df.iloc[-count:,0] = model_names
        df.iloc[-count:,3]= xgbs_df.iloc[h_start:h_end,2].values
        combined_df_store.append(df.sort_values('Val RMSE'))

    wb = openpyxl.Workbook()
    for h, df in zip([1, 3, 6, 12, 24], combined_df_store):
        wb.create_sheet(f'h_{h}')
        ws = wb[f'h_{h}']
        print_df_to_excel(df=df, ws=ws)

    wb.save(f'{results_dir}/best summary + xgbs.xlsx')