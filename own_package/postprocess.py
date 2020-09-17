import numpy as np
import pandas as pd
import math, random
import cvxpy as cp
import openpyxl, pickle, collections
import matplotlib.pyplot as plt
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from own_package.dm_test import dm_test
from own_package.others import create_excel_file, create_results_directory


def difference_to_levels(save_dir_store, h_store, rawdata_excel, first_est_date, varname):
    df_master = pd.read_excel(rawdata_excel, sheet_name='Master')
    transformation_type = df_master[varname].iloc[0]
    df_master.drop([0], inplace=True)
    df_master.index = [f'{x}:{y}' for x, y in
                           zip(pd.DatetimeIndex(df_master['sasdate']).year, pd.DatetimeIndex(df_master['sasdate']).day)]

    for save_dir,h_idx, h in zip(save_dir_store,range(len(save_dir_store)), h_store):
        with open(save_dir, 'rb') as handle:
            data_store = pickle.load(handle)
        data_df = data_store['data_df']
        df = data_df[[x for x in data_df.columns if '_ehat' in x]].copy()
        df = data_df[[x for x in data_df.columns if 'y_' in x]].values - df
        df = pd.concat((df_master[[varname]], df), axis=1)
        start_idx = np.where(df.index==first_est_date)[0][0]
        if transformation_type == 5:
            levelhat = df[[varname]].iloc[start_idx-int(h):-int(h)].values*np.exp(int(h)/1200*df[[x for x in data_df.columns if '_ehat' in x]].iloc[start_idx:])
        elif transformation_type == 6:
            xt = df[[varname]].iloc[start_idx-int(h):-int(h)].values
            xt_1 = df[[varname]].iloc[start_idx-int(h)-1:-int(h)-1].values
            levelhat = xt ** (int(h)+1) / (xt_1 ** int(h)) * np.exp(int(h)/1200*df[[x for x in data_df.columns if '_ehat' in x]].iloc[start_idx:])
        else:
            raise TypeError(f'Invalid transformation type: {transformation_type}')
        data_df[f'y_{h}'] = df_master[[varname]].iloc[start_idx:]
        data_df[[x for x in data_df.columns if '_ehat' in x]] = df_master[[varname]].iloc[start_idx:].values - levelhat
        data_store['data_df'] = data_df
        with open(save_dir.partition('.pkl')[0] + '_levels.pkl', 'wb') as handle:
            pickle.dump(data_store, handle)


def combine_poos_analysis(results_dir, dir_store, levels=False, combined_name='combined'):
    if levels:
        levels = '_levels'
    else:
        levels = ''

    h_store = [1,3,6,12,24]

    for h in h_store:
        df_store = []
        for dir in dir_store:
            data_df = None
            for file in os.listdir(dir):
                if f'h{h}_analysis_results{levels}.pkl' in file:
                    with open(f'{dir}/{file}', 'rb') as handle:
                        data_df = pickle.load(handle)['data_df']
                    data_df = data_df[[f'y_{h}']+[x for x in data_df.columns if '_ehat' in x]]
                    data_df[[x for x in data_df.columns if '_ehat' in x]] = data_df[[f'y_{h}']].values-data_df[[x for x in data_df.columns if '_ehat' in x]]
                    df_store.append(data_df.copy())

            if data_df is None:
                raise FileNotFoundError(f'For {dir}, h step = {h} analysis results pickle file was not found')

        # Take simple avg. of all the yhat from the different models in dir_store
        data_df = pd.concat(objs=df_store, axis=0).groupby(level=0).mean()
        data_df[[x for x in data_df.columns if '_ehat' in x]] = data_df[[f'y_{h}']].values - data_df[
            [x for x in data_df.columns if '_ehat' in x]]

        with open(f'{results_dir}/poos_{combined_name}_h{h}_analysis_results{levels}.pkl', 'wb') as handle:
            pickle.dump({'data_df': data_df, 'combined_dir_store':dir_store}, handle)


def plot_forecasts(save_dir_store, results_dir, model_names, est_store, h_store):
    results_dir = create_results_directory(results_dir)
    for h, sds in zip(h_store, save_dir_store):
        for idx, (model_name, est, save_dir) in enumerate(zip(model_names, est_store, sds)):
            with open(save_dir, 'rb') as handle:
                data_df = pickle.load(handle)['data_df']

            if idx == 0:
                df = data_df[[f'y_{h}']]

            if model_name in ['ar', 'pca']:
                df = pd.concat((df, data_df[[f'{model_name}_ehat']]), axis=1)
            elif model_name == 'xgba':
                df = pd.concat((df, data_df[[f'rw_ehat']].rename(columns={'rw_ehat':f'xgba_rw_{est}_ehat'}, inplace=False)), axis=1)
            elif model_name == 'rf':
                df = pd.concat((df, data_df[[f'rf_ehat']].rename(columns={'rf_ehat':f'rf_{est}_ehat'}, inplace=False)), axis=1)


        ax = df[[x for x in df.columns if '_ehat' in x]].plot(lw=0.5)
        ax.ylabel = 'ehat'
        plt.savefig(f'{results_dir}/{h}_ehat_all.png')
        plt.close()

        ax = df[[x for x in df.columns if any([y in x for y in ['ar', 'pca', 'rw_rh']])]].plot(lw=0.5)
        ax.ylabel = 'ehat'
        plt.savefig(f'{results_dir}/{h}_ehat_arpcaxgbarwrh.png')
        plt.close()



    pass


def read_excel_to_df(excel_dir):
    xls = pd.ExcelFile(excel_dir)
    sheet_names = xls.sheet_names
    df_store = []
    for sheet in sheet_names:
        if sheet == 'Sheet':
            pass
        else:
            try:
                df = pd.read_excel(excel_dir, sheet_name=sheet, skiprows=[0, 2],
                                   index_col=0).sort_values(['m', 'p'])
            except KeyError:
                # Best summary does not have that blank line
                df = pd.read_excel(excel_dir, sheet_name=sheet,
                                   index_col=0)
            df = df.reset_index(drop=True)
            df_store.append(df)

    return df_store


def compile_pm_rm_excel(excel_dir_store):
    master_pm = [[] for x in range(5)]
    master_rm = [[] for x in range(5)]
    for excel_dir in excel_dir_store:
        xls = pd.ExcelFile(excel_dir)
        sheet_names = xls.sheet_names[1:]
        for sheet, pm_store, rm_store in zip(sheet_names, master_pm, master_rm):
            df = pd.read_excel(excel_dir, sheet_name=sheet, index_col=None).values
            pm_store.append(df[1:10, :])
            rm_store.append(df[11:, 0][..., None])

    for idx, pm_h in enumerate(master_pm):
        pm = pm_h[0]
        for pm_hh in pm_h[1:]:
            pm = np.concatenate((pm, pm_hh), axis=1)
        master_pm[idx] = pm

    for idx, pm_h in enumerate(master_rm):
        rm = pm_h[0]
        for pm_hh in pm_h[1:]:
            rm = np.concatenate((rm, pm_hh), axis=1)
        master_rm[idx] = rm

    excel_dir = create_excel_file('./results/master_pm_rd.xlsx')
    wb = openpyxl.load_workbook(excel_dir)
    for idx, (pm, rm) in enumerate(zip(master_pm, master_rm)):

        pm_name = 'pm_h{}'.format([1, 3, 6, 12, 24][idx])
        rm_name = 'rm_h{}'.format([1, 3, 6, 12, 24][idx])
        wb.create_sheet(pm_name)
        wb.create_sheet(rm_name)

        ws = wb[pm_name]
        pm_df = pd.DataFrame(data=pm, columns=['m', 'p'] * len(excel_dir_store))
        rows = dataframe_to_rows(pm_df, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx + 1, column=c_idx, value=value)

        ws = wb[rm_name]
        rm_df = pd.DataFrame(data=rm, columns=['Relative RMSE'] * len(excel_dir_store))
        rows = dataframe_to_rows(rm_df, index=False)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx + 1, column=c_idx, value=value)

    wb.save(excel_dir)

    pass


class Postdata:
    def __init__(self, results_dir, var_name, calculations=True, star=True):
        self.star = star
        self.results_dir = results_dir
        # First 3 lines if a list of dataframes. Each df is one h step ahead, for h=1,3,6,12,24
        # 4th line is list of 1D ndarray for y values
        # 5th line is list of 2D ndarray for (models, y hat values)

        # self.testset_XXX_y is the list of  1d ndarray for the XXX model (e.g. PCA) actual y values
        # self.testset_XXX_y_hat is the list of  2d ndarray for the XXX model, size = (no. of models, no. of y values)
        # self.testset_XXX_AWA_y_hat is the list of 1d ndarray for the XXX model forecasted y values combined with AWA
        # self.testset_XXX_BWA_y_hat is the list of 1d ndarray for the XXX model forecasted y values combined with BWA
        self.AR_AIC_BIC = read_excel_to_df('{}/{}_AR_AIC_BIC.xlsx'.format(results_dir, var_name))
        self.AR_PLS = read_excel_to_df('{}/{}_AR_PLS.xlsx'.format(results_dir, var_name))
        self.testset_AR_PLS = read_excel_to_df('{}/testset_{}_AR_PLS.xlsx'.format(results_dir, var_name))
        self.testset_AR_y = [np.array(x.columns.tolist()[5:]) for x in self.testset_AR_PLS]
        self.testset_AR_y_hat = [x.iloc[:, 5:].values for x in self.testset_AR_PLS]

        self.PCA_AIC_BIC = read_excel_to_df('{}/{}_PCA_AIC_BIC.xlsx'.format(results_dir, var_name))
        self.PCA_PLS = read_excel_to_df('{}/{}_PCA_PLS.xlsx'.format(results_dir, var_name))
        self.testset_PCA_PLS = read_excel_to_df('{}/testset_{}_PCA_PLS.xlsx'.format(results_dir, var_name))
        self.testset_PCA_y = [np.array(x.columns.tolist()[5:]) for x in self.testset_PCA_PLS]
        self.testset_PCA_y_hat = [x.iloc[:, 5:].values for x in self.testset_PCA_PLS]

        self.UMAP_AIC_BIC = read_excel_to_df('{}/{}_UMAP_AIC_BIC.xlsx'.format(results_dir, var_name))
        self.UMAP_PLS = read_excel_to_df('{}/{}_UMAP_PLS.xlsx'.format(results_dir, var_name))
        self.testset_UMAP_PLS = read_excel_to_df('{}/testset_{}_UMAP_PLS.xlsx'.format(results_dir, var_name))
        self.testset_UMAP_y = [np.array(x.columns.tolist()[5:]) for x in self.testset_UMAP_PLS]
        self.testset_UMAP_y_hat = [x.iloc[:, 5:].values for x in self.testset_UMAP_PLS]

        self.hsteps = [1, 3, 6, 12, 24]

        if calculations:
            self.num_h = len(self.AR_AIC_BIC)
            self.pm_store = [np.zeros((9, 2)) for x in range(self.num_h)]
            self.rm_store = [np.zeros((33), dtype=np.object) for x in range(self.num_h)]
            self.benchmark_rmse = []
            self.benchmarky = []

            i = 0
            # Iterate through each h step ahead values for all AR. h = 1,3,6,12,24
            for idx, (aic, pls, test, pm, rm, yhat, y) in enumerate(
                    zip(self.AR_AIC_BIC, self.AR_PLS, self.testset_AR_PLS, self.pm_store, self.rm_store,
                        self.testset_AR_y_hat, self.testset_AR_y)):
                self.benchmark_forecasted_y_BIC = []
                min_BIC_idx = np.argmin(aic['BIC_t'])
                pm[1, 1] = aic['p'][min_BIC_idx]
                rmse_idx = test.index[test['p'] == pm[1, 1]].tolist()[0]
                self.benchmark_rmse.append(test['Val RMSE'][rmse_idx])
                self.benchmarky.append(yhat[rmse_idx])
                rm[1] = round(test['Val RMSE'][rmse_idx], 4)



                min_AIC_idx = np.argmin(aic['AIC_t'])
                pm[0, 1] = aic['p'][min_AIC_idx]
                rmse_idx2 = test.index[test['p'] == pm[0, 1]].tolist()[0]
                rmse = test['Val RMSE'][rmse_idx2]
                rm[0] = round(rmse / self.benchmark_rmse[-1], 4)
                forecastedy = yhat[rmse_idx2]
                if rmse_idx != rmse_idx2:

                    dm_r = dm_test(y, self.benchmarky[i], forecastedy, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[0] = '{}*'.format(round(rm[0], 4))



                min_idx = np.argmin(pls['Val RMSE'])
                pm[2, 1] = pls['p'][min_idx]
                rmse_idx2 = test.index[test['p'] == pm[2, 1]].tolist()[0]
                rmse = test['Val RMSE'][rmse_idx2]
                rm[2] = round(rmse / self.benchmark_rmse[-1], 4)
                forecastedy = yhat[rmse_idx2]
                if rmse_idx != rmse_idx2:

                    dm_r = dm_test(y, self.benchmarky[i], forecastedy, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[2] = '{}*'.format(round(rm[2], 4))


                i = i + 1

            i = 0
            # Iterate through each h step ahead values for all PCA. h = 1,3,6,12,24
            skip = 3
            skip2 = 8
            for idx, (aic, pls, test, pm, rm, yhat, y) in enumerate(
                    zip(self.PCA_AIC_BIC, self.PCA_PLS, self.testset_PCA_PLS, self.pm_store, self.rm_store,
                        self.testset_PCA_y_hat, self.testset_PCA_y)):
                min_BIC_idx = np.argmin(aic['BIC_t'])
                pm[1 + skip, 0] = aic['m'][min_BIC_idx]
                pm[1 + skip, 1] = aic['p'][min_BIC_idx]
                rmse_idx = test.index[(test['m'] == pm[1 + skip, 0]) & (test['p'] == pm[1 + skip, 1])].tolist()[0]
                rmse = test['Val RMSE'][rmse_idx]
                rm[1 + skip2] = round(rmse / self.benchmark_rmse[idx], 4)
                forecastedy = yhat[rmse_idx]
                if np.all(self.benchmarky[i] != forecastedy):
                    dm_r = dm_test(y, self.benchmarky[i], forecastedy, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[1 + skip2] = '{}*'.format(rm[1 + skip2])


                min_AIC_idx = np.argmin(aic['AIC_t'])
                pm[0 + skip, 0] = aic['m'][min_AIC_idx]
                pm[0 + skip, 1] = aic['p'][min_AIC_idx]
                rmse_idx = test.index[(test['m'] == pm[0 + skip, 0]) & (test['p'] == pm[0 + skip, 1])].tolist()[0]
                rmse = test['Val RMSE'][rmse_idx]
                rm[0 + skip2] = round(rmse / self.benchmark_rmse[idx], 4)
                forecastedy = yhat[rmse_idx]
                if np.all(self.benchmarky[i] != forecastedy):
                    dm_r = dm_test(y, self.benchmarky[i], forecastedy, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[0 + skip2] = '{}*'.format(round(rm[0 + skip2], 4))


                min_idx = np.argmin(pls['Val RMSE'])
                pm[2 + skip, 0] = pls['m'][min_idx]
                pm[2 + skip, 1] = pls['p'][min_idx]
                rmse_idx = test.index[(test['m'] == pm[2 + skip, 0]) & (test['p'] == pm[2 + skip, 1])].tolist()[0]
                rmse = test['Val RMSE'][rmse_idx]
                rm[2 + skip2] = round(rmse / self.benchmark_rmse[idx], 4)
                forecastedy = yhat[rmse_idx]
                if np.all(self.benchmarky[i] != forecastedy):
                    dm_r = dm_test(y, self.benchmarky[i], forecastedy, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[2 + skip2] = '{}*'.format(round(rm[2 + skip2], 4))


                i = i + 1

            i = 0
            # Iterate through each h step ahead values for all UMAP. h = 1,3,6,12,24
            skip = 3 * 2
            skip2 = 8 * 2
            for idx, (aic, pls, test, pm, rm, yhat, y) in enumerate(
                    zip(self.UMAP_AIC_BIC, self.UMAP_PLS, self.testset_UMAP_PLS, self.pm_store, self.rm_store,
                        self.testset_UMAP_y_hat, self.testset_UMAP_y)):
                min_BIC_idx = np.argmin(aic['BIC_t'])
                pm[1 + skip, 0] = aic['m'][min_BIC_idx]
                pm[1 + skip, 1] = aic['p'][min_BIC_idx]
                rmse_idx = test.index[(test['m'] == pm[1 + skip, 0]) & (test['p'] == pm[1 + skip, 1])].tolist()[0]
                rmse = test['Val RMSE'][rmse_idx]
                rm[1 + skip2] = round(rmse / self.benchmark_rmse[idx], 4)
                forecastedy = yhat[rmse_idx]
                if np.all(self.benchmarky[i] != forecastedy):
                    dm_r = dm_test(y, self.benchmarky[i], forecastedy, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[1 + skip2] = '{}*'.format(round(rm[1 + skip2], 4))


                min_AIC_idx = np.argmin(aic['AIC_t'])
                pm[0 + skip, 0] = aic['m'][min_AIC_idx]
                pm[0 + skip, 1] = aic['p'][min_AIC_idx]
                rmse_idx = test.index[(test['m'] == pm[0 + skip, 0]) & (test['p'] == pm[0 + skip, 1])].tolist()[0]
                rmse = test['Val RMSE'][rmse_idx]
                rm[0 + skip2] = round(rmse / self.benchmark_rmse[idx], 4)
                forecastedy = yhat[rmse_idx]
                if np.all(self.benchmarky[i] != forecastedy):
                    dm_r = dm_test(y, self.benchmarky[i], forecastedy, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[0 + skip2] = '{}*'.format(round(rm[0 + skip2], 4))


                min_idx = np.argmin(pls['Val RMSE'])
                pm[2 + skip, 0] = pls['m'][min_idx]
                pm[2 + skip, 1] = pls['p'][min_idx]
                rmse_idx = test.index[(test['m'] == pm[2 + skip, 0]) & (test['p'] == pm[2 + skip, 1])].tolist()[0]
                rmse = test['Val RMSE'][rmse_idx]
                rm[2 + skip2] = round(rmse / self.benchmark_rmse[idx], 4)
                forecastedy = yhat[rmse_idx]
                if np.all(self.benchmarky[i] != forecastedy):
                    dm_r = dm_test(y, self.benchmarky[i], forecastedy, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[2 + skip2] = '{}*'.format(round(rm[2 + skip2], 4))


                i = i + 1

            pass

    def combination(self):
        """

        :param type: Either 'AIC_t' or 'BIC_t' for AWA and BWA respectively
        :return:
        """
        aic_bic_store = [self.AR_AIC_BIC, self.PCA_AIC_BIC, self.UMAP_AIC_BIC]
        pls_store = [self.AR_PLS, self.PCA_PLS, self.UMAP_PLS]
        testset_y_store = [self.testset_AR_y, self.testset_PCA_y, self.testset_UMAP_y]
        testset_y_hat_store = [self.testset_AR_y_hat, self.testset_PCA_y_hat, self.testset_UMAP_y_hat]
        self.testset_AR_AWA_y_hat = []
        self.testset_AR_BWA_y_hat = []
        self.testset_AR_AVG_y_hat = []
        self.testset_AR_GR_y_hat = []
        self.testset_PCA_AWA_y_hat = []
        self.testset_PCA_BWA_y_hat = []
        self.testset_PCA_AVG_y_hat = []
        self.testset_PCA_GR_y_hat = []
        self.testset_UMAP_AWA_y_hat = []
        self.testset_UMAP_BWA_y_hat = []
        self.testset_UMAP_AVG_y_hat = []
        self.testset_UMAP_GR_y_hat = []
        self.testset_PU_AVG_y_hat = []
        self.testset_PU_GR_y_hat = []

        for skip_idx, (aic_bic_all_h, pls_all_h, testset_y, testset_y_hat, awa_y_hat, bwa_y_hat, avg_y_hat, gr_y_hat) \
                in enumerate(zip(aic_bic_store, pls_store, testset_y_store, testset_y_hat_store,
                                 [self.testset_AR_AWA_y_hat, self.testset_PCA_AWA_y_hat, self.testset_UMAP_AWA_y_hat],
                                 [self.testset_AR_BWA_y_hat, self.testset_PCA_BWA_y_hat, self.testset_UMAP_BWA_y_hat],
                                 [self.testset_AR_AVG_y_hat, self.testset_PCA_AVG_y_hat, self.testset_UMAP_AVG_y_hat],
                                 [self.testset_AR_GR_y_hat, self.testset_PCA_GR_y_hat, self.testset_UMAP_GR_y_hat])):
            i = 0
            for idx, (ic, pls, y, y_hat, rm) in enumerate(
                    zip(aic_bic_all_h, pls_all_h, testset_y, testset_y_hat, self.rm_store)):
                # Simple average AVG
                t_idx = 3 + 8 * skip_idx
                y_combi_hat = np.mean(y_hat, axis=0)
                avg_y_hat.append(y_combi_hat)
                rmse_combi = math.sqrt(np.mean(np.array(y - y_combi_hat) ** 2))
                rm[t_idx] = round(rmse_combi / self.benchmark_rmse[idx], 4)
                if np.all(self.benchmarky[i] != y_combi_hat):
                    dm_r = dm_test(y, self.benchmarky[i], y_combi_hat, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[t_idx] = '{}*'.format(round(rm[t_idx], 4))


                # AWA
                type = 'AIC_t'
                t_idx = 4 + 8 * skip_idx
                ic_values = ic[type].values
                min_ic = np.min(ic_values)
                ic_values += -min_ic
                weights = np.exp(-ic_values / 2)
                weights = weights / np.sum(weights)
                y_combi_hat = np.sum(y_hat * weights[:, None], axis=0)
                awa_y_hat.append(y_combi_hat)
                rmse_combi = math.sqrt(np.mean(np.array(y - y_combi_hat) ** 2))
                rm[t_idx] = round(rmse_combi / self.benchmark_rmse[idx], 4)
                if np.all(self.benchmarky[i] != y_combi_hat):
                    dm_r = dm_test(y, self.benchmarky[i], y_combi_hat, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[t_idx] = '{}*'.format(round(rm[t_idx], 4))


                # BWA
                type = 'BIC_t'
                t_idx = 5 + 8 * skip_idx
                ic_values = ic[type].values
                min_ic = np.min(ic_values)
                ic_values += -min_ic
                weights = np.exp(-ic_values / 2)
                weights = weights / np.sum(weights)
                y_combi_hat = np.sum(y_hat * weights[:, None], axis=0)
                bwa_y_hat.append(y_combi_hat)
                rmse_combi = math.sqrt(np.mean(np.array(y - y_combi_hat) ** 2))
                rm[t_idx] = round(rmse_combi / self.benchmark_rmse[idx], 4)
                if np.all(self.benchmarky[i] != y_combi_hat):
                    dm_r = dm_test(y, self.benchmarky[i], y_combi_hat, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[t_idx] = '{}*'.format(round(rm[t_idx], 4))


                # GR
                t_idx = 6 + 8 * skip_idx
                y_pls = np.array(pls.columns.tolist()[5:])
                y_hat_pls = pls.iloc[:, 5:].values
                #m = np.shape(y_hat_pls)[0] + 1  # number of models + 1 constant term
                m = np.shape(y_hat_pls)[0]  # number of models + 1 constant term
                n = np.shape(y_hat_pls)[1]  # number of timesteps
                beta = cp.Variable(shape=(m, 1))

                # pc_1 = np.ones((1, m - 1)) @ beta[1:, 0] == 1
                pc_1 = np.ones((1, m)) @ beta == 1
                pc_2 = beta >= 0
                constraints = [pc_1, pc_2]

                # X = np.concatenate((np.ones((n, 1)), y_hat_pls.T), axis=1)
                X = y_hat_pls.T
                z = np.ones((1, n)) @ (y_pls[:, None] - X @ beta) ** 2
                objective = cp.Minimize(z)
                prob = cp.Problem(objective, constraints)

                prob.solve(solver='GUROBI')
                beta_hat = beta.value
                '''
                print('Skip_idx: {} idx: {} sum beta: {:.3E} min beta: {:.3E} max beta: {:.3E}'.format(skip_idx, idx,
                                                                                                       np.sum(beta_hat),
                                                                                                       np.min(beta_hat),
                                                                                                       np.max(beta_hat)))
                
                print('Skip_idx: {} idx: {} sum beta: {:.3E} min beta: {:.3E} max beta: {:.3E}'.format(skip_idx, idx,
                                                                                                       np.sum(beta_hat[
                                                                                                              1:]),
                                                                                                       np.min(beta_hat[
                                                                                                              1:]),
                                                                                                       np.max(beta_hat[
                                                                                                              1:])))
                '''
                y_combi_hat = np.sum(y_hat * beta_hat[:, 0][:, None], axis=0)
                gr_y_hat.append(y_combi_hat)
                rmse_combi = math.sqrt(np.mean(np.array(y - y_combi_hat) ** 2))
                rm[t_idx] = round(rmse_combi / self.benchmark_rmse[idx], 4)
                if np.all(self.benchmarky[i] != y_combi_hat):
                    dm_r = dm_test(y, self.benchmarky[i], y_combi_hat, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[t_idx] = '{}*'.format(round(rm[t_idx], 4))

                # GR with intercept
                t_idx = 7 + 8 * skip_idx
                y_pls = np.array(pls.columns.tolist()[5:])
                y_hat_pls = pls.iloc[:, 5:].values
                m = np.shape(y_hat_pls)[0] + 1  # number of models + 1 constant term
                #m = np.shape(y_hat_pls)[0]  # number of models + 1 constant term
                n = np.shape(y_hat_pls)[1]  # number of timesteps
                beta = cp.Variable(shape=(m, 1))

                pc_1 = np.ones((1, m - 1)) @ beta[1:, 0] == 1
                # pc_1 = np.ones((1, m)) @ beta == 1
                pc_2 = beta >= 0
                constraints = [pc_1, pc_2]

                X = np.concatenate((np.ones((n, 1)), y_hat_pls.T), axis=1)
                # X = y_hat_pls.T
                z = np.ones((1, n)) @ (y_pls[:, None] - X @ beta) ** 2
                objective = cp.Minimize(z)
                prob = cp.Problem(objective, constraints)

                prob.solve(solver='GUROBI')
                beta_hat = beta.value
                '''
                print('Skip_idx: {} idx: {} sum beta: {:.3E} min beta: {:.3E} max beta: {:.3E}'.format(skip_idx, idx,
                                                                                                       np.sum(beta_hat),
                                                                                                       np.min(beta_hat),
                                                                                                       np.max(beta_hat)))

                print('Skip_idx: {} idx: {} sum beta: {:.3E} min beta: {:.3E} max beta: {:.3E}'.format(skip_idx, idx,
                                                                                                       np.sum(beta_hat[
                                                                                                              1:]),
                                                                                                       np.min(beta_hat[
                                                                                                              1:]),
                                                                                                       np.max(beta_hat[
                                                                                                              1:])))
                '''
                y_combi_hat = np.sum(y_hat * beta_hat[1:, 0][:, None] + beta_hat[0,0], axis=0)
                gr_y_hat.append(y_combi_hat)
                rmse_combi = math.sqrt(np.mean(np.array(y - y_combi_hat) ** 2))
                rm[t_idx] = round(rmse_combi / self.benchmark_rmse[idx], 4)
                if np.all(self.benchmarky[i] != y_combi_hat):
                    dm_r = dm_test(y, self.benchmarky[i], y_combi_hat, h=self.hsteps[i], crit="MSE")
                    pvalue = dm_r[1]
                    if pvalue <= 0.05 and self.star:
                        rm[t_idx] = '{}*'.format(round(rm[t_idx], 4))


                i = i + 1
        i = 0
        # PCA+UMAP
        for idx, (pca_pls, umap_pls, y, pca_y_hat, umap_y_hat, rm) in enumerate(
                zip(self.PCA_PLS, self.UMAP_PLS, self.testset_PCA_y, self.testset_PCA_y_hat, self.testset_UMAP_y_hat,
                    self.rm_store)):
            # AVG
            y_pls = np.array(pca_pls.columns.tolist()[5:])
            pca_y_hat_pls = pca_pls.iloc[:, 5:].values
            umap_y_hat_pls = umap_pls.iloc[:, 5:].values

            y_combi_hat = np.mean(np.concatenate((pca_y_hat, umap_y_hat), axis=0), axis=0)
            self.testset_PU_AVG_y_hat.append(y_combi_hat)
            rmse_combi = math.sqrt(np.mean(np.array(y - y_combi_hat) ** 2))
            rm[24] = round(rmse_combi / self.benchmark_rmse[idx], 4)
            if np.all(self.benchmarky[i] != y_combi_hat):
                dm_r = dm_test(y, self.benchmarky[i], y_combi_hat, h=self.hsteps[i], crit="MSE")
                pvalue = dm_r[1]
                if pvalue <= 0.05 and self.star:
                    rm[24] = '{}*'.format(round(rm[24], 4))


            # GR
            y_hat_pls = np.concatenate((pca_y_hat_pls, umap_y_hat_pls), axis=0)
            #m = np.shape(y_hat_pls)[0] + 1  # number of models + 1 constant term
            m = np.shape(y_hat_pls)[0]
            n = np.shape(y_hat_pls)[1]  # number of timesteps
            beta = cp.Variable(shape=(m, 1))

            #pc_1 = np.ones((1, m - 1)) @ beta[1:, 0] == 1
            pc_1 = np.ones((1, m)) @ beta== 1
            pc_2 = beta >= 0
            constraints = [pc_1, pc_2]

            #X = np.concatenate((np.ones((n, 1)), y_hat_pls.T), axis=1)
            X = y_hat_pls.T

            z = np.ones((1, n)) @ (y_pls[:, None] - X @ beta) ** 2
            objective = cp.Minimize(z)
            prob = cp.Problem(objective, constraints)

            prob.solve(solver='GUROBI')
            beta_hat = beta.value
            '''
            print('idx: {} sum beta: {:.3E} min beta: {:.3E} max beta: {:.3E}'.format( idx,
                                                                                                   np.sum(beta_hat),
                                                                                                   np.min(beta_hat),
                                                                                                   np.max(beta_hat)))
            
            print('idx: {} sum beta: {:.3E} min beta: {:.3E} max beta: {:.3E}'.format(idx,
                                                                                      np.sum(beta_hat[1:]),
                                                                                      np.min(beta_hat[1:]),
                                                                                      np.max(
                                                                                          beta_hat[1:])))
            '''
            y_hat = np.concatenate((pca_y_hat, umap_y_hat), axis=0)
            y_combi_hat = np.sum(y_hat * beta_hat[:, 0][:, None] , axis=0)
            self.testset_PU_GR_y_hat.append(y_combi_hat)
            rmse_combi = math.sqrt(np.mean(np.array(y - y_combi_hat) ** 2))
            rm[25] = round(rmse_combi / self.benchmark_rmse[idx], 4)
            if np.all(self.benchmarky[i] != y_combi_hat):
                dm_r = dm_test(y, self.benchmarky[i], y_combi_hat, h=self.hsteps[i], crit="MSE")
                pvalue = dm_r[1]
                if pvalue <= 0.05 and self.star:
                    rm[25] = '{}*'.format(round(rm[25], 4))

            # GR with intercept
            y_hat_pls = np.concatenate((pca_y_hat_pls, umap_y_hat_pls), axis=0)
            m = np.shape(y_hat_pls)[0] + 1  # number of models + 1 constant term
            # m = np.shape(y_hat_pls)[0]
            n = np.shape(y_hat_pls)[1]  # number of timesteps
            beta = cp.Variable(shape=(m, 1))

            pc_1 = np.ones((1, m - 1)) @ beta[1:, 0] == 1
            #pc_1 = np.ones((1, m)) @ beta == 1
            pc_2 = beta >= 0
            constraints = [pc_1, pc_2]

            X = np.concatenate((np.ones((n, 1)), y_hat_pls.T), axis=1)
            #X = y_hat_pls.T

            z = np.ones((1, n)) @ (y_pls[:, None] - X @ beta) ** 2
            objective = cp.Minimize(z)
            prob = cp.Problem(objective, constraints)

            prob.solve(solver='GUROBI')
            beta_hat = beta.value
            '''
            print('idx: {} sum beta: {:.3E} min beta: {:.3E} max beta: {:.3E}'.format(idx,
                                                                                      np.sum(beta_hat),
                                                                                      np.min(beta_hat),
                                                                                      np.max(beta_hat)))
            
            print('idx: {} sum beta: {:.3E} min beta: {:.3E} max beta: {:.3E}'.format(idx,
                                                                                      np.sum(beta_hat[1:]),
                                                                                      np.min(beta_hat[1:]),
                                                                                      np.max(
                                                                                          beta_hat[1:])))
            '''
            y_hat = np.concatenate((pca_y_hat, umap_y_hat), axis=0)
            y_combi_hat = np.sum(y_hat * beta_hat[1:, 0][:, None]+ beta_hat[0,0], axis=0)
            self.testset_PU_GR_y_hat.append(y_combi_hat)
            rmse_combi = math.sqrt(np.mean(np.array(y - y_combi_hat) ** 2))
            rm[26] = round(rmse_combi / self.benchmark_rmse[idx], 4)
            if np.all(self.benchmarky[i] != y_combi_hat):
                dm_r = dm_test(y, self.benchmarky[i], y_combi_hat, h=self.hsteps[i], crit="MSE")
                pvalue = dm_r[1]
                if pvalue <= 0.05 and self.star:
                    rm[26] = '{}*'.format(round(rm[26], 4))

            i = i + 1

        i = 0
        # decomp_combi
        def run_decomp_combi(subgroup_size, numel, rm_idx):
            all_h_y_hat = [np.array(ar.tolist() + pca.tolist() + umap.tolist()) for ar, pca, umap in
                           zip(self.testset_AR_y_hat, self.testset_PCA_y_hat, self.testset_UMAP_y_hat)]
            model_count = [single_all_y_hat.shape[0] for single_all_y_hat in all_h_y_hat]
            selections = [random.sample(list(range(model_count[0])), k=subgroup_size) for _ in range(numel)]
            for idx, (single_all_y_hat, single_y, h_label, rm) in enumerate(zip(all_h_y_hat, self.testset_AR_y, self.hsteps, self.rm_store)):
                # perform sub selection for each h step ahead
                sub_y_hat_store = np.array([single_all_y_hat[selection, :] for selection in selections])
                sub_y_mean_hat = np.mean(sub_y_hat_store, axis=1)
                sub_y_invvar_hat = np.reciprocal(np.var(sub_y_hat_store, axis=1))
                total_weights = np.sum(sub_y_invvar_hat, axis=0)
                p_y = np.sum((1 / total_weights * sub_y_mean_hat * sub_y_invvar_hat), axis=0)
                rm[rm_idx] = round(np.sqrt(np.average(np.square(p_y - single_y))) / self.benchmark_rmse[idx], 4)

        subgroup_size = 20
        numel = 50
        rm_idx = 27
        run_decomp_combi(subgroup_size=subgroup_size, numel=numel, rm_idx=rm_idx)

        subgroup_size = 20
        numel = 500
        rm_idx = 28
        run_decomp_combi(subgroup_size=subgroup_size, numel=numel, rm_idx=rm_idx)

        subgroup_size = 20
        numel = 5000
        rm_idx = 29
        run_decomp_combi(subgroup_size=subgroup_size, numel=numel, rm_idx=rm_idx)

        subgroup_size = 10
        numel = 50
        rm_idx = 30
        run_decomp_combi(subgroup_size=subgroup_size, numel=numel, rm_idx=rm_idx)

        subgroup_size = 10
        numel = 500
        rm_idx = 31
        run_decomp_combi(subgroup_size=subgroup_size, numel=numel, rm_idx=rm_idx)

        subgroup_size = 10
        numel = 5000
        rm_idx = 32
        run_decomp_combi(subgroup_size=subgroup_size, numel=numel, rm_idx=rm_idx)


        # Printing to excel
        excel_dir = create_excel_file('{}/pm_rm_results.xlsx'.format(self.results_dir))
        wb = openpyxl.load_workbook(excel_dir)
        for idx in range(len(self.pm_store)):
            wb.create_sheet('h = {}'.format([1, 3, 6, 12, 24][idx]))
        sheet_names = wb.sheetnames

        for sheet, pm, rm in zip(sheet_names[1:], self.pm_store, self.rm_store):
            ws = wb[sheet]

            pm_df = pd.DataFrame(data=pm, columns=['m', 'p'])
            rows = dataframe_to_rows(pm_df, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx + 1, column=c_idx, value=value)

            skip = len(pm_df.index) + 1
            rm_df = pd.DataFrame(rm, columns=['Relative RMSE'])
            rows = dataframe_to_rows(rm_df, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx + 1 + skip, column=c_idx, value=value)

        wb.save(excel_dir)

        pass
