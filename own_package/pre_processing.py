from own_package.features_labels import read_excel_data, Fl_master, Fl_pca
from own_package.others import create_results_directory, print_df_to_excel, create_excel_file
import numpy as np
import pandas as pd
import openpyxl, math, collections
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from statsmodels.multivariate.pca import PCA as SMPCA
from statsmodels.tsa.stattools import adfuller, kpss
from arch.unitroot import DFGLS
import pymannkendall as mk
import matplotlib.pyplot as plt


def type_transformations(excel_dir, results_dir, y_selection, h_steps):
    df = pd.read_excel(excel_dir, sheet_name='Master')
    names = df.columns.values.tolist()
    data = df.values
    data_type_store = np.copy(data[0, 1:])
    time_stamps = np.copy(data[3:, 0])
    data = np.copy(data[1:, 1:]).astype(np.float)

    x_store = []
    for _, (type, x) in enumerate(zip(data_type_store.tolist(), data.T.tolist())):
        if type == 1:
            x_store.append(x)
        elif type == 2:
            x_transformed = np.array(x)[1:] - np.array(x)[:-1]
            x_transformed = [np.nan] + x_transformed.tolist()
            x_store.append(x_transformed)
        elif type == 4:
            x_transformed = np.log(np.array(x)).tolist()
            x_store.append(x_transformed)
        elif type == 5:
            x_transformed = np.log(np.array(x)[1:]) - np.log(np.array(x)[:-1])
            x_transformed = [np.nan] + x_transformed.tolist()
            x_store.append(x_transformed)
        elif type == 6:
            x_transformed = np.log(np.array(x)[2:]) - 2 * np.log(np.array(x)[1:-1]) + np.log(np.array(x)[:-2])
            x_transformed = [np.nan, np.nan] + x_transformed.tolist()
            x_store.append(x_transformed)
        elif type == 7:
            x_transformed = np.array(x)[2:] / np.array(x)[1:-1] - np.array(x)[1:-1] / np.array(x)[:-2]
            x_transformed = [np.nan, np.nan] + x_transformed.tolist()
            x_store.append(x_transformed)
        else:
            pass

    x_store = np.array(x_store).T

    temp_names = names[1:]
    selection_idx = [i for i in range(len(temp_names)) if temp_names[i] in y_selection]

    y_transformed_names = []
    y_store = []
    for idx, selection in enumerate(selection_idx):
        yo = data[:, selection]
        type = data_type_store[selection]
        for h in h_steps:
            y_transformed_names.append('{}_h{}'.format(temp_names[selection], h))
            if type == 5:
                y_transformed = 1200 / h * np.log(yo[h:] / yo[:-h])
                y_transformed = [np.nan] * h + y_transformed.tolist()
                y_store.append(y_transformed)
            elif type == 6:
                y_transformed = 1200 / h * np.log(yo[h + 1:] / yo[1:-h]) - 1200 * np.log(yo[1:-h] / yo[:-h - 1])
                y_transformed = [np.nan] * (h + 1) + y_transformed.tolist()
                y_store.append(y_transformed)
            else:
                raise KeyError('Label type is not 5 or 6')

    y_store = (np.array(y_store).T)[2:, :]
    x_store[:, selection_idx] = x_store[:, selection_idx] * 1200
    x_store = x_store[2:, :]

    # _, ic, v = iterated_em(all_x=x_store.copy(), pca_p=9, max_iter=1e4, tol=0.1)
    pc = SMPCA(data=x_store.copy(), ncomp=9, missing='fill-em')
    x_store = pc._adjusted_data

    results_dir = create_results_directory(results_dir)
    wb = openpyxl.Workbook()
    wb.create_sheet('transformation')
    sheet_name = wb.sheetnames[-1]
    ws = wb[sheet_name]
    df = pd.DataFrame(data=np.concatenate((time_stamps[..., None], x_store), axis=1),
                      columns=names)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    wb.create_sheet('y transformed')
    sheet_name = wb.sheetnames[-1]
    ws = wb[sheet_name]
    ydf = pd.DataFrame(data=np.concatenate((time_stamps[..., None], y_store), axis=1),
                      columns=['Time Stamps'] + y_transformed_names)
    for r in dataframe_to_rows(ydf, index=False, header=True):
        ws.append(r)

    def summary_test(df, data_type_store):
        results_dict = collections.defaultdict(dict)
        suggested_type_store = []
        for var, type_ in zip(df.columns.values[1:], data_type_store):
            ts = df[var].values.astype(float)
            # ADF test. Null: time series has a unit root
            adf_p = adfuller(x=ts.copy())[1]
            # KPSS test. Null: time series is stationary around a constant
            kpss_p = kpss(x=ts.copy())[1]
            results_dict[var]['adf p_value'] = adf_p
            results_dict[var]['kpss p_value'] = kpss_p
            '''
            Case 1: Both tests conclude that the series is not stationary - The series is not stationary
            Case 2: Both tests conclude that the series is stationary - The series is stationary
            Case 3: KPSS indicates stationarity and ADF indicates non-stationarity - 
            The series is trend stationary. Trend needs to be removed to make series strict stationary. 
            The detrended series is checked for stationarity.
            Case 4: KPSS indicates non-stationarity and ADF indicates stationarity - 
            The series is difference stationary. Differencing is to be used to make series stationary. 
            The differenced series is checked for stationarity.
            '''
            if adf_p >= 0.05 and kpss_p <= 0.05:
                case = 1
                suggested_type_store.append(type_+1)  # Try differencing
            elif adf_p <= 0.05 and kpss_p >= 0.05:
                case = 2
                suggested_type_store.append(type_)
            elif adf_p>=0.05 and kpss_p>=0.05:
                case = 3
                suggested_type_store.append('BAD THERE IS TREND')
            elif adf_p<=0.05 and kpss_p<=0.05:
                case = 4
                suggested_type_store.append(type_ + 1)  # Try differencing
            results_dict[var]['stationary case'] = case

            # DF-GLS test
            dfgls = DFGLS(ts.copy())
            results_dict[var]['dfgls p_value'] = dfgls.pvalue

            # Yue wang modified Mann-Kendall tests account for serial autocorrelation. Null: No monotonic trend
            mk_test = mk.yue_wang_modification_test(ts.copy())
            results_dict[var]['ywmk p_value'] = mk_test.p
            results_dict[var]['ywmk slope'] = mk_test.slope
            results_dict[var]['ywmk trend'] = mk_test.trend
        ret = pd.DataFrame.from_dict(results_dict, orient='index').reset_index()
        ret['type'] = data_type_store
        ret['suggested type'] = suggested_type_store
        count_non2 = (ret['stationary case'] != 2).sum()
        count_trend = (ret['ywmk trend'] != 'no trend').sum()
        ret['type'] = data_type_store

        ret = pd.concat((pd.DataFrame.from_dict({'fail count':{'stationary case':count_non2, 'ywmk trend': count_trend}}, orient='index'), ret), axis=0)
        ret = ret[['index', 'type', 'suggested type', 'adf p_value', 'kpss p_value', 'stationary case', 'dfgls p_value',
                   'ywmk p_value', 'ywmk slope', 'ywmk trend']]
        return ret

    time_stamps = [f'{x}:{y}' for x, y in
                   zip(pd.DatetimeIndex(time_stamps).year, pd.DatetimeIndex(time_stamps).day)]

    # Stationary tests for transformed X
    wb.create_sheet('tests')
    ws = wb['tests']
    print_df_to_excel(df=summary_test(df.copy(), data_type_store=data_type_store), ws=ws)
    df['sasdate'] = time_stamps
    for var in df.columns.values[1:]:
        plt.close()
        df.plot(y=var, x='sasdate')
        plt.savefig(f'{results_dir}/plot_{var}.png')

    # Stationary tests for transformed X BUT exclude COVID ==> Last period 2019:12
    wb.create_sheet('tests 2019-12')
    ws = wb['tests 2019-12']
    print_df_to_excel(df=summary_test(df.iloc[:-6, :].copy(), data_type_store=data_type_store), ws=ws)

    idx = np.where(np.array(time_stamps) == '2015:12')[0][0]
    wb.create_sheet('tests 2015-12')
    ws = wb['tests 2015-12']
    print_df_to_excel(df=summary_test(df.iloc[:idx+1, :].copy(), data_type_store=data_type_store), ws=ws)

    wb.create_sheet('em fill ic')
    ws = wb['em fill ic']
    ic_store = collections.defaultdict(dict)
    for stop_date in ['2020:6'] + [f'{year}:12' for year in (2019 - np.arange(50))]:
        for r in range(1, 13):
            idx = np.where(np.array(time_stamps) == stop_date)[0][0]
            ic = SMPCA(data=x_store.copy()[:idx + 1, :], ncomp=r, missing='fill-em').ic[-1, 1].copy()
            ic_store[time_stamps[idx]][r] = ic
    df = pd.DataFrame.from_dict(ic_store)
    df.index.name = 'k'
    mini = df.min()
    idxmin = df.idxmin()
    df.loc['min'] = mini
    df.loc['best k'] = idxmin
    print_df_to_excel(df=df, ws=ws)

    wb.save('{}/transformed_data.xlsx'.format(results_dir))
    create_data_loader_excel(excel_dir='{}/transformed_data.xlsx'.format(results_dir), results_dir=results_dir)


def pca_factor_estimation(x, r, N, x_transformed_already=False):
    if not x_transformed_already:
        x_scaler = StandardScaler()
        x_scaler.fit(x)
        x = x_scaler.transform(x)

    # w, v = eigh(x.T @ x)
    # loadings = np.fliplr(v[:, -r:])
    # loadings = loadings * math.sqrt(N)

    pca = PCA(n_components=r, svd_solver='full')
    pca.fit(x)
    loadings = pca.components_.T * math.sqrt(N)
    factors = x @ loadings / N
    loadings_T = loadings.T

    return factors, loadings_T


def iterated_em(all_x, pca_p, max_iter, tol):
    N = all_x.shape[1]
    nan_store = np.where(np.isnan(all_x))
    col_mean = np.nanmean(all_x, axis=0)
    all_x[nan_store] = np.take(col_mean, nan_store[1])

    diff = 1000
    iter = 0
    while iter < max_iter:
        if diff < tol:
            print('Tolerance met with maximum percentage error = {}%\n Number of iterations taken = {}'.format(
                diff, iter))
            break
        all_x_scaler = StandardScaler()
        all_x_scaler.fit(all_x)
        all_x_norm = all_x_scaler.transform(all_x)

        factors, loadings_T = pca_factor_estimation(x=all_x_norm, r=pca_p, N=N, x_transformed_already=True)

        all_x_norm_1 = factors @ loadings_T
        all_x_1 = all_x_scaler.inverse_transform(all_x_norm_1)

        try:
            # diff = max((all_x_1[nan_store] - all_x[nan_store]) / all_x[nan_store] * 100)  # Tolerance is in percentage error
            diff = np.max(np.abs((factors - factors_old) / factors_old * 100))
            factors_old = factors
        except:
            diff = 1000
            factors_old = factors
        all_x[nan_store] = all_x_1[nan_store]
        iter += 1
    else:
        raise ValueError(
            'During iterated EM method, maximum iteration of {} without meeting tolerance requirement.'
            ' Current maximum percentage error = {}%'.format(iter, diff))

    # IC score calculation
    nobs = all_x.shape[0]
    x_hat = all_x_scaler.inverse_transform(factors @ loadings_T)
    v = 1 / (nobs * N) * np.sum((all_x - x_hat) ** 2)
    k = np.shape(factors)[1]

    # Using g2 penalty.
    c_nt = min(nobs, N)

    return all_x, math.log(v) + k * ((nobs + N) / (nobs * N)) * math.log(c_nt), v


def create_data_loader_excel(excel_dir, results_dir):
    ymain_df = pd.read_excel(excel_dir, sheet_name='y transformed', index_col=0)
    xmain_df = pd.read_excel(excel_dir, 'transformation', index_col=0)

    # Find unique var name for forecasting
    var_names = list(set([item.partition('_h')[0] for item in ymain_df.columns]))

    for var_name in var_names:
        excel_name = create_excel_file('{}/{}_data_loader.xlsx'.format(results_dir, var_name))
        wb = openpyxl.load_workbook(excel_name)
        wb.create_sheet('x')
        wb.create_sheet('yo')
        wb.create_sheet('y')
        print_df_to_excel(df=xmain_df.loc[:, xmain_df.columns != var_name], ws=wb['x'])
        print_df_to_excel(df=xmain_df.loc[:, [var_name]], ws=wb['yo'])
        mask = np.flatnonzero(np.core.defchararray.find(ymain_df.columns.values.astype(str), var_name) != -1)
        print_df_to_excel(df=ymain_df.iloc[:, mask], ws=wb['y'])
        wb.save(excel_name)

    pass


# Fractional Differencing Part from Advances in Financial Machine Learning
def getWeights(d,size):
    # thres>0 drops insignificant weights
    w=[1.]
    for k in range(1,size):
        w_=-w[-1]/k*(d-k+1)
        w.append(w_)
    w=np.array(w[::-1]).reshape(-1,1)
    return w


def fracDiff_FFD(series,d,thres=1e-5):
    '''
    Constant width window (new solution)
    Note 1: thres determines the cut-off weight for the window
    Note 2: d can be any positive fractional, not necessarily bounded [0,1].
    '''
    #1) Compute weights for the longest series
    w=getWeights(d,thres)
    width=len(w)-1
    #2) Apply weights to values
    df={}
    for name in series.columns:
        seriesF, df_ = series[[name]].fillna(method='ffill').dropna(), pd.Series()
        for iloc1 in range(width, seriesF.shape[0]):
            loc0, loc1 = seriesF.index[iloc1 - width], seriesF.index[iloc1]
            if not np.isfinite(series.loc[loc1, name]):
                continue  # exclude NAs
            df_[loc1] = np.dot(w.T, seriesF.loc[loc0:loc1])[0, 0]
        df[name] = df_.copy(deep=True)
    df = pd.concat(df, axis=1)
    return df

