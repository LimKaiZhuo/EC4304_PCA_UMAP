import numpy as np
from scipy.linalg import fractional_matrix_power
from scipy.stats import zscore
import pandas as pd
import statsmodels.api as sm
import openpyxl
import matplotlib.pyplot as plt
from matplotlib import cm
import seaborn as sns
import joypy
from arch.bootstrap import MCS, SPA
from own_package.boosting import Xgboost
from own_package.ssm import LocalLevel, SSMBase
from own_package.others import print_df_to_excel, create_excel_file, create_results_directory, set_matplotlib_style
import pickle, time, itertools, shap, collections, fnmatch

first_est_date = '1970:1'


def forecast_error(y_predicted, y_true):
    return 'ehat', np.sum(y_true.get_label() - y_predicted)


def poos_experiment(fl_master, fl, est_dates, z_type, h, h_idx, m_max, p_max, first_est_date,
                    model_mode, save_dir,
                    **kwargs):
    est_dates = est_dates + [fl_master.time_stamp[-1]]

    data_store = []

    for idx, (est_date, next_tune_date) in enumerate(zip(est_dates[:-1], est_dates[1:])):
        t1 = time.perf_counter()
        (x_est, _), (yo_est, _), (y_est, _), (ts_est, _), _, _ = fl_master.date_split(est_date)
        (x_tt, _), (yo_tt, _), (y_tt, _), (ts_tt, _), _, _ = fl_master.date_split(next_tune_date, date_start=est_date)
        if model_mode == 'xgb':
            hparams_df = fl.xgb_hparam_opt(x=x_est, yo=yo_est, y=y_est[:, [h_idx]], h=h,
                                           m_max=m_max, p_max=p_max,
                                           z_type=z_type,
                                           hparam_opt_params=kwargs['hparam_opt_params'],
                                           default_hparams=kwargs['default_hparams'],
                                           results_dir=None,
                                           model_name=None)

            # wb = openpyxl.Workbook()
            # ws = wb[wb.sheetnames[-1]]
            # print_df_to_excel(df=hparams_df, ws=ws)
            # wb.save('./results/poos/hparams.xlsx')
            # print('done')

            hparams = {**kwargs['default_hparams'], **hparams_df.iloc[0, :].to_dict()}
            hparams['early_stopping_rounds'] = None
            hparams['m'] = int(hparams['m'])
            hparams['max_depth'] = int(hparams['max_depth'])
            hparams['ehat_eval'] = forecast_error
            _, _, _, poos_data_store = fl.pls_expanding_window(h=h, p=hparams['m'] * 2, m=hparams['m'], r=9,
                                                               cw_model_class=Xgboost,
                                                               cw_hparams=hparams,
                                                               x_t=x_est,
                                                               x_v=x_tt,
                                                               yo_t=yo_est,
                                                               y_t=y_est[:, h_idx][..., None],
                                                               yo_v=yo_tt,
                                                               y_v=y_tt[:, h_idx][..., None],
                                                               rolling=True,
                                                               z_type=z_type,
                                                               save_dir=None,
                                                               save_name=None)
            data_store.append({'est_date': est_date,
                               'next_tune_date': next_tune_date,
                               'hparams_df': hparams_df,
                               'poos_data_store': poos_data_store},
                              )
            with open('{}/poos_h{}.pkl'.format(save_dir, h), "wb") as file:
                pickle.dump(data_store, file)
        elif model_mode == 'xgb_with_hparam':
            hparams = {**kwargs['default_hparams'], **kwargs['set_hparam'][h]}
            hparams['early_stopping_rounds'] = None
            hparams['m'] = int(hparams['m'])
            hparams['max_depth'] = int(hparams['max_depth'])
            hparams['ehat_eval'] = forecast_error
            _, _, _, poos_data_store = fl.pls_expanding_window(h=h, p=hparams['m'] * 2, m=hparams['m'], r=8,
                                                               cw_model_class=Xgboost,
                                                               cw_hparams=hparams,
                                                               x_t=x_est,
                                                               x_v=x_tt,
                                                               yo_t=yo_est,
                                                               y_t=y_est[:, h_idx][..., None],
                                                               yo_v=yo_tt,
                                                               y_v=y_tt[:, h_idx][..., None],
                                                               rolling=True,
                                                               z_type=z_type,
                                                               save_dir=None,
                                                               save_name=None)
            data_store.append({'est_date': est_date,
                               'next_tune_date': next_tune_date,
                               'hparams_df': pd.DataFrame.from_dict({0:kwargs['set_hparam'][h]}, orient='index'),
                               'poos_data_store': poos_data_store},
                              )
            with open('{}/poos_h{}.pkl'.format(save_dir, h), "wb") as file:
                pickle.dump(data_store, file)
        elif model_mode == 'rf':
            hparams_df = fl.xgb_hparam_opt(x=x_est, yo=yo_est, y=y_est[:, [h_idx]], h=h,
                                           m_max=m_max, p_max=p_max,
                                           z_type=z_type,
                                           hparam_opt_params=kwargs['hparam_opt_params'],
                                           default_hparams=kwargs['default_hparams'],
                                           results_dir=None,
                                           model_name=None)

            # wb = openpyxl.Workbook()
            # ws = wb[wb.sheetnames[-1]]
            # print_df_to_excel(df=hparams_df, ws=ws)
            # wb.save('./results/poos/hparams.xlsx')
            # print('done')

            hparams = {**kwargs['default_hparams'], **hparams_df.iloc[0, :].to_dict()}
            hparams['early_stopping_rounds'] = None
            hparams['m'] = int(hparams['m'])
            hparams['max_depth'] = int(hparams['max_depth'])
            hparams['num_parallel_tree'] = int(hparams['num_parallel_tree'])
            hparams['ehat_eval'] = forecast_error
            _, _, _, poos_data_store = fl.pls_expanding_window(h=h, p=hparams['m'] * 2, m=hparams['m'], r=8,
                                                               cw_model_class=Xgboost,
                                                               cw_hparams=hparams,
                                                               x_t=x_est,
                                                               x_v=x_tt,
                                                               yo_t=yo_est,
                                                               y_t=y_est[:, h_idx][..., None],
                                                               yo_v=yo_tt,
                                                               y_v=y_tt[:, h_idx][..., None],
                                                               rolling=True,
                                                               z_type=z_type,
                                                               save_dir=None,
                                                               save_name=None)
            data_store.append({'est_date': est_date,
                               'next_tune_date': next_tune_date,
                               'hparams_df': hparams_df,
                               'poos_data_store': poos_data_store},
                              )
            with open('{}/poos_h{}.pkl'.format(save_dir, h), "wb") as file:
                pickle.dump(data_store, file)
        elif model_mode == 'pca':
            hparams_df = fl.pca_hparam_opt(x=x_est, yo=yo_est, y=y_est[:, [h_idx]], h=h,
                                           m_max=m_max, p_max=p_max, r=8)
            hparams = hparams_df.iloc[0, :]
            yhat, ehat, _, _ = fl.pls_expanding_window(h=h, p=int(hparams['p']), m=int(hparams['m']), r=8,
                                                       factor_model=fl.pca_factor_estimation,
                                                       x_t=x_est,
                                                       x_v=x_tt,
                                                       yo_t=yo_est,
                                                       y_t=y_est[:, h_idx][..., None],
                                                       yo_v=yo_tt,
                                                       y_v=y_tt[:, h_idx][..., None],
                                                       rolling=True,
                                                       save_dir=None,
                                                       save_name=None)
            data_store.append({'est_date': est_date,
                               'next_tune_date': next_tune_date,
                               'hparams_df': hparams_df,
                               'ehat': ehat,
                               'yhat': yhat},
                              )
        elif model_mode == 'ar':
            hparams_df = fl.ar_hparam_opt(yo=yo_est, y=y_est[:, [h_idx]], h=h, p_max=p_max, )
            hparams = hparams_df.iloc[0, :]
            yhat, ehat, _, _ = fl.ar_pls_expanding_window(h=h, p=int(hparams['p']),
                                                          yo_t=yo_est,
                                                          y_t=y_est[:, h_idx][..., None],
                                                          yo_v=yo_tt,
                                                          y_v=y_tt[:, h_idx][..., None], rolling=True,
                                                          save_dir=None, save_name=None)
            data_store.append({'est_date': est_date,
                               'next_tune_date': next_tune_date,
                               'hparams_df': hparams_df,
                               'ehat': ehat,
                               'yhat': yhat},
                              )

        elif model_mode == 'rw':
            yhat, ehat, _ = fl.rw_pls_expanding_window(y_t=y_est[:, h_idx][..., None], y_v=y_tt[:, h_idx][..., None],
                                                       h=h, save_dir=None, save_name=None)
            data_store.append({'est_date': est_date,
                               'next_tune_date': next_tune_date,
                               'ehat': ehat,
                               'yhat': yhat},
                              )

        t2 = time.perf_counter()
        print(f'Completed {est_date} for h={h}. Time taken {t2 - t1}')

    if model_mode == 'ar4':
        _, (yo_est, yo_tt), (y_est, y_tt), (ts_est, ts_tt), _, _ = fl_master.date_split(est_dates[0])
        _, e_hat_store, _, _ = fl.ar_pls_expanding_window(h=h, p=4,
                                                          yo_t=yo_est,
                                                          y_t=y_est[:, h_idx][..., None],
                                                          yo_v=yo_tt,
                                                          y_v=y_tt[:, h_idx][..., None], rolling=True,
                                                          save_dir=None, save_name=None)

        idx = fl_master.time_stamp.index(first_est_date)
        y = fl_master.y[idx:, h_idx]
        ts = fl_master.time_stamp[idx:]

        data_df = pd.DataFrame.from_dict({'time_stamp': ts,
                                          f'y_{h}': y,
                                          'ar4_ehat': e_hat_store, }).set_index('time_stamp')
        hparam_df = pd.DataFrame(data=np.array([4] * len(est_dates[:-1]))[..., None], columns=['p'])
        with open(f'{save_dir}/poos_{model_mode}_h{h}_analysis_results.pkl', "wb") as file:
            pickle.dump({'data_df': data_df, 'hparam_df': hparam_df}, file)

    elif model_mode in ['pca', 'ar', 'rw']:
        try:
            hparam_df = pd.concat([data['hparams_df'].iloc[0, :] for data in data_store], axis=1).T
        except KeyError:
            # Means it is rw as rw has no hyperparameters
            hparam_df = None

        idx = fl_master.time_stamp.index(first_est_date)
        y = fl_master.y[idx:, h_idx]
        ts = fl_master.time_stamp[idx:]

        data_df = pd.DataFrame.from_dict({'time_stamp': ts,
                                          f'y_{h}': y,
                                          f'{model_mode}_ehat': [x for data in data_store for x in data['ehat']],
                                          }).set_index('time_stamp')
        with open(f'{save_dir}/poos_{model_mode}_h{h}_analysis_results.pkl', "wb") as file:
            pickle.dump({'data_df': data_df, 'hparam_df': hparam_df}, file)


def poos_analysis(fl_master, h, h_idx, model_mode, est_mode, results_dir, save_dir, first_est_date):
    def calculate_sequence_ntree(block_ae, hparam_ntree):
        best_idx_store = list(np.argmin(block_ae, axis=1))
        predicted_idx = [hparam_ntree] * 5
        current_window = best_idx_store[:5]
        for y in best_idx_store[5:]:
            model = LocalLevel(current_window)
            res = model.fit(disp=False)
            predicted_idx.append(int(res.forecast()))
            current_window.append(y)
        return best_idx_store, predicted_idx

    def calculate_ssm_ntree(hparam_ntree, optimal_ntree, mode, burn_periods=10):
        best_idx_store = optimal_ntree
        predicted_idx = [hparam_ntree] + optimal_ntree[:(burn_periods - 1)]
        current_window = best_idx_store[:burn_periods]
        for y in best_idx_store[burn_periods:]:
            base = SSMBase(mode=mode)
            base.fit(X=current_window)
            predicted_idx.append(int(max(min(base.predict(), 599), 0)))
            current_window.append(y)
        return predicted_idx, base.res_

    def calculate_horizon_rmse_ae(block_ae, block_ehat, sequence_ntree):
        return np.average([ae[idx] ** 2 for ae, idx in zip(block_ae, sequence_ntree)]) ** 0.5, \
               [ehat[idx] for ehat, idx in zip(block_ehat, sequence_ntree)]

    def calculate_hparam_ntree(hparam_df):
        if est_mode == 'rh':
            div_factor = 0.85
        elif est_mode == 'rfcv':
            div_factor = 0.8
        return min(int(round(hparam_df.iloc[0, hparam_df.columns.tolist().index('m iters')] / div_factor)), 600) - 1

    with open(save_dir, 'rb') as handle:
        data_store = pickle.load(handle)

    idx = fl_master.time_stamp.index(first_est_date)
    y = fl_master.y[idx:, h_idx]
    ts = fl_master.time_stamp[idx:]

    if model_mode == 'xgb':
        index_products = [('y', str(idx)) for idx in range(3)]
        score_df = pd.DataFrame(index=pd.MultiIndex.from_tuples(index_products, names=['Variable', 'Lag'])).T
        ae_store = []
        ehat_store = []
        optimal_ntree_store = []
        predicted_ntree_store = []
        rw_ntree_store = []
        hparam_ntree_store = []
        block_store = []
        ssm_modes = ['arima*ln', 'll', 'lls', 'llt', 'llc', 'll*ln', 'lls*ln', 'llt*ln', 'llc*ln', ]
        ssm_store = {}
        for k in ssm_modes:
            ssm_store[k] = {f'{k}_ntree': [], f'{k}_ehat': [], f'{k}_res': []}

        for idx, data in enumerate(data_store):
            for single_step in data['poos_data_store']:
                scores = {(k.partition('_L')[0], k.partition('_L')[-1]): v for k, v in
                          single_step['feature_score'].items()}
                score_df = score_df.append(scores, ignore_index=True)

            block_ae = [single_step['progress']['h_step_ahead']['rmse'] for single_step in data['poos_data_store']]
            block_store.extend([idx] * len(block_ae))
            block_ehat = [single_step['progress']['h_step_ahead']['ehat'] for single_step in data['poos_data_store']]
            ae_store.extend(block_ae)
            ehat_store.extend(block_ehat)

            hparam_ntree = calculate_hparam_ntree(data['hparams_df'])
            optimal_ntree = list(np.argmin(block_ae, axis=1))
            optimal_ntree_store.extend(optimal_ntree)

            for k, v in ssm_store.items():
                ntree, res = calculate_ssm_ntree(hparam_ntree, optimal_ntree, mode=k)
                v[f'{k}_ntree'].extend(ntree)
                v[f'{k}_res'].append(res)

            # _, predicted_ntree = calculate_sequence_ntree(block_ae=block_ae,
            #                                                           hparam_ntree=hparam_ntree)
            # predicted_ntree_store.extend(predicted_ntree)

            rw_ntree_store.extend([hparam_ntree] + optimal_ntree[:-1])
            hparam_ntree_store.extend([hparam_ntree] * len(optimal_ntree))

        # Full SSM prediction
        # _, ssm_full_ntree_store = calculate_sequence_ntree(block_ae=ae_store, hparam_ntree=hparam_ntree_store[0])
        ll_full_ntree, ll_full_res = calculate_ssm_ntree(hparam_ntree=hparam_ntree_store[0],
                                                         optimal_ntree=optimal_ntree_store,
                                                         mode='ll*ln', burn_periods=24)
        # Horizons RMSE calculation
        _, ll_full_ehat = calculate_horizon_rmse_ae(ae_store, ehat_store, ll_full_ntree)
        oracle_rmse, oracle_ehat = calculate_horizon_rmse_ae(ae_store, ehat_store, optimal_ntree_store)
        max_iter_rmse, max_iter_ehat = calculate_horizon_rmse_ae(ae_store, ehat_store, [-1] * len(ae_store))
        hparam_rmse, hparam_ehat = calculate_horizon_rmse_ae(ae_store, ehat_store, hparam_ntree_store)
        # _, ssm_ehat = calculate_horizon_rmse_ae(ae_store, ehat_store, predicted_ntree_store)
        # ssm_full_rmse, ssm_full_ehat = calculate_horizon_rmse_ae(ae_store, ehat_store, ssm_full_ntree_store)
        rw_rmse, rw_ehat = calculate_horizon_rmse_ae(ae_store, ehat_store, rw_ntree_store)

        for k, v in ssm_store.items():
            _, ehat = calculate_horizon_rmse_ae(ae_store, ehat_store, v[f'{k}_ntree'])
            v[f'{k}_ehat'] = ehat

        df = pd.DataFrame.from_dict({**{'time_stamp': ts,
                                        'hparam_block': block_store,
                                        f'y_{h}': y,
                                        'oracle_ehat': oracle_ehat,
                                        'll_full_ehat': ll_full_ehat,
                                        # 'ssm_ehat': ssm_ehat,
                                        # 'ssm_full_ehat': ssm_full_ehat,
                                        'rw_ehat': rw_ehat,
                                        'hparam_ehat': hparam_ehat,
                                        'max_iter_ehat': max_iter_ehat,
                                        'oracle_ntree': optimal_ntree_store,
                                        # 'ssm_full_ntree': ssm_full_ntree_store,
                                        'll_full_ntree': ll_full_ntree,
                                        'rw_ntree': rw_ntree_store,
                                        'hparam_ntree': hparam_ntree_store},
                                     **{kk: vv for _, v in ssm_store.items() for kk, vv in v.items() if
                                        'ntree' in kk or 'ehat' in kk}}).set_index(
            'time_stamp')

        res_store = {k: v[f'{k}_res'] for k, v in ssm_store.items()}

        hparam_df = [data['hparams_df'].iloc[0, :] for data in data_store]
        hparam_df = pd.concat(hparam_df, axis=1).T

        with open(f'{results_dir}/poos_{model_mode}_h{h}_analysis_results.pkl', "wb") as file:
            pickle.dump({'data_df': df, 'hparam_df': hparam_df, 'res_store': res_store}, file)
    elif model_mode == 'rf':
        df = pd.DataFrame.from_dict({'time_stamp': ts,
                                     f'y_{h}': y,
                                     'rf_ehat': [step['progress']['h_step_ahead']['ehat'][0] for block in data_store for
                                                 step in
                                                 block['poos_data_store']], }).set_index('time_stamp')
        hparam_df = data_store[0]['hparams_df'].iloc[[0], :]

        with open(f'{results_dir}/poos_{model_mode}_h{h}_analysis_results.pkl', "wb") as file:
            pickle.dump({'data_df': df, 'hparam_df': hparam_df}, file)


def poos_analysis_combining_xgb(h, results_dir, poos_post_dir_store):
    yhat_df_store = []
    hparam_df_store = []
    for poos_post_dir in poos_post_dir_store:
        with open(f'{poos_post_dir}/poos_xgb_h{h}_analysis_results.pkl', 'rb') as handle:
            data = pickle.load(handle)
        df = data['data_df']
        yhat_df = df[[x for x in df.columns if 'y_' in x]].values - df[[x for x in df.columns if '_ehat' in x]]
        yhat_df_store.append(yhat_df)
        hparam_df_store.append(data['hparam_df'].reset_index(drop=True))

    y = df[[x for x in df.columns if 'y_' in x]].values
    data_df = pd.concat(yhat_df_store).groupby(level=0).mean().reindex(df.index)
    data_df = y - data_df
    hparam_df = pd.concat(hparam_df_store).groupby(level=0).mean()

    with open(f'{results_dir}/poos_xgb_h{h}_analysis_results.pkl', "wb") as file:
        pickle.dump({'data_df': data_df, 'hparam_df': hparam_df}, file)


def poos_xgb_plotting_m(h, results_dir, ssm_modes):
    import matplotlib as mpl
    try:
        del mpl.font_manager.weight_dict['roman']
        mpl.font_manager._rebuild()
    except KeyError:
        pass
    sns.set(style='ticks')
    mpl.rc('font', family='Times New Roman')

    with open(f'{results_dir}/poos_xgb_h{h}_analysis_results.pkl', 'rb') as handle:
        data = pickle.load(handle)

    def plots(df, names, plot_name):
        names = [f'{x}_ntree' for x in names]
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.margins(x=0.01, y=0.01)
        df.plot(kind='line', y=['oracle_ntree'] + names, ax=ax, marker='o', ms=1.5, lw=0.5)
        ax.set_ylabel('ntrees')
        fig.tight_layout()
        fig.savefig(f'{results_dir}/{plot_name}_ntrees_vs_time_h{h}.png', transparent=False, dpi=300,
                    bbox_inches="tight")
        plt.close()

        temp = df[names].transform(lambda x: np.log(x + 1))
        temp['oracle_ntree'] = df['oracle_ntree'].transform(lambda x: np.log(x + 1))
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.margins(x=0.01, y=0.01)
        temp.plot(kind='line', y=['oracle_ntree'] + names, ax=ax, marker='o', ms=1.5, lw=0.5)
        ax.set_ylabel('ln(ntrees+1)')
        fig.tight_layout()
        fig.savefig(f'{results_dir}/{plot_name}_ln_ntrees_vs_time_h{h}.png', transparent=False, dpi=300,
                    bbox_inches="tight")
        plt.close()

    # Plot ntree vs time
    df = data['data_df']
    plots(df, ['ll', 'll*ln'], plot_name='ll')
    plots(df, ['llt', 'llt*ln'], plot_name='llt')
    plots(df, ['lls', 'lls*ln'], plot_name='lls')
    plots(df, ['llc', 'llc*ln'], plot_name='llc')

    # Plot last block residual diagnostics
    for k, v in data['res_store'].items():
        if not 'arima' in k:
            plt.close()
            v[-1].plot_diagnostics()
            name = k.replace('*', '_')
            plt.tight_layout(rect=[0, 0.03, 1, 0.95])  # to make subtitles not overlap with x axis label
            plt.savefig(f'{results_dir}/{name}_residual_diagnostics_h{h}.png')
    plt.close()

    with open(f'{results_dir}/poos_h{h}.pkl', 'rb') as handle:
        data = pickle.load(handle)
    ehat_store = np.array(
        [step['progress']['h_step_ahead']['rmse'] for block in data for step in block['poos_data_store']])
    ehat_store = (ehat_store - ehat_store.min(axis=1)[:, None]) / (
            ehat_store.max(axis=1)[:, None] - ehat_store.min(axis=1)[:, None])
    ehat_df = pd.DataFrame(ehat_store.T, columns=df.index.values)
    fig, ax = plt.subplots()
    ax.margins(x=0.01, y=0.01)
    try:
        y_store = ['1970:1', '1983:1', '2007:1', '2020:3']
        ehat_df.plot(kind='line', y=y_store, ax=ax, lw=0.5)
    except KeyError:
        y_store = ['2005:1', '2005:2', '2010:1', '2010:2']
        ehat_df.plot(kind='line', y=y_store, ax=ax, lw=0.7)

    ax.set_ylabel('|ehat|')
    ax.set_xlabel('ntrees')
    ax.set_xlim([0, 300])
    ax.legend(loc='right')
    for y in y_store:
        ax.scatter(df.loc[y][f'{ssm_modes[0]}_ntree'], ehat_df[y].iloc[int(df.loc[y][f'{ssm_modes[0]}_ntree'])],
                   marker='x', s=70)
    fig.tight_layout()
    fig.savefig(f'{results_dir}/{ssm_modes[0].replace("*", "_")}ehat_vs_ntrees_h{h}.png', transparent=False, dpi=300,
                bbox_inches="tight")


def poos_processed_data_analysis(results_dir, save_dir_store, h_store, model_full_name, nber_excel_dir, first_est_date,
                                 est_dates, model, combinations=None):
    if first_est_date == '2005:1':
        # Means it is expt 1. Skip NBER and structural RMSE windows
        skip = True
    else:
        skip = False
        # Load NBER dates
        nber = pd.read_excel(nber_excel_dir, sheet_name='main')
        peaks = pd.DatetimeIndex(pd.to_datetime(nber['Peak month']))[1:]  # First month is NaN
        troughs = pd.DatetimeIndex(pd.to_datetime(nber['Trough month']))[1:]
        dates = pd.date_range('1970-01-01', '2020-12-01', freq='MS')
        nber_df = pd.DataFrame(data=[-1] * len(dates), index=dates, columns=['business_cycle'])
        for idx, (d1, d2) in enumerate(zip(peaks, troughs)):
            # recession ==> 0
            nber_df.loc[d1:d2 - pd.offsets.DateOffset(1, 'months')] = 0
        for idx, (d1, d2) in enumerate(zip(troughs[:-1], peaks[1:])):
            # expansionary ==> 1
            nber_df.loc[d1:d2 - pd.offsets.DateOffset(1, 'months')] = 1
        nber_df.index = [f'{int(year)}:{int(month)}' for year, month in
                         zip(nber_df.index.year.values, nber_df.index.month.values)]

    est_dates = est_dates[1:]
    df_store = []
    for save_dir in save_dir_store:
        with open(save_dir, 'rb') as handle:
            data_store = pickle.load(handle)

        data_df = data_store['data_df']
        if model == 'xgb':
            if combinations:
                combinations_store = []
                for combi in combinations:
                    combi_name = '+'.join(combi) + '_ehat'
                    combinations_store.append(
                        pd.DataFrame(data_df[[x for x in data_df.columns if 'y_' in x]].values.squeeze() -
                                     (data_df[[x for x in data_df.columns if 'y_' in x]].values
                                      - data_df[[f'{x}_ehat' for x in combi]]).mean(axis=1),
                                     columns=[combi_name]))
                data_df = pd.concat([data_df] + combinations_store, axis=1)

        ehat_df = data_df[[x for x in data_df.columns.values if '_ehat' in x]]
        ehat_df.columns = [name.replace('ehat', 'rmse') for name in ehat_df.columns.values]
        time_stamps = data_df.index.tolist()
        est_dates_idx = [0] + [time_stamps.index(est) + 1 for est in est_dates] + [-1]

        rmse_store = []
        for start, end in zip(est_dates_idx[:-1], est_dates_idx[1:]):
            rmse_store.append((ehat_df.iloc[start:end, :] ** 2).mean(axis=0) ** 0.5)

        if not skip:
            # Means include the categorical horizons
            start_store = ['1970:1', '1970:1', '1970:1', '1984:4', '2007:1', '2019:12']  # Inclusive
            end_store = ['2020:1', '2020:12', '1984:4', '2007:1', '2020:1', '2020:12']  # Exclusive of those dates

            for start, end in zip(start_store, end_store):
                start = time_stamps.index(start)
                try:
                    end = time_stamps.index(end)
                    rmse_store.append((ehat_df.iloc[start:end, :] ** 2).mean(axis=0) ** 0.5)
                except ValueError:
                    # Means end is the last date
                    rmse_store.append((ehat_df.iloc[start:, :] ** 2).mean(axis=0) ** 0.5)

            # NBER Expansionary and Recessionary
            rmse_store.append((ehat_df.loc[nber_df['business_cycle'] == 1, :] ** 2).mean(axis=0) ** 0.5)
            rmse_store.append((ehat_df.loc[nber_df['business_cycle'] == 0, :] ** 2).mean(axis=0) ** 0.5)
        else:
            start_store = ['2005:1', '2020:1']  # Inclusive
            end_store = ['2020:1', '2020:12']  # Exclusive of those dates

            for start, end in zip(start_store, end_store):
                start = time_stamps.index(start)
                try:
                    end = time_stamps.index(end)
                    rmse_store.append((ehat_df.iloc[start:end, :] ** 2).mean(axis=0) ** 0.5)
                except ValueError:
                    # Means end is the last date
                    rmse_store.append((ehat_df.iloc[start:, :] ** 2).mean(axis=0) ** 0.5)

        # Combining data into dataframe
        try:
            df = pd.concat((pd.concat(rmse_store, axis=1).T, data_store['hparam_df'].reset_index()), axis=1)
        except (KeyError, AttributeError):
            # AttributeError means it is rw which does not have hparam_df
            df = pd.concat(rmse_store, axis=1).T

        if not skip:
            df['start_dates'] = [f'{x}:1' for x in range(1970, 2020, 5)] + start_store + ['Expansionary',
                                                                                          'Recessionary']
            df['end_dates'] = [f'{x}:1' for x in range(1970, 2020, 5)[1:]] + [time_stamps[-1]] + end_store + ['', '']
        else:
            df['start_dates'] = [first_est_date] + est_dates + start_store
            df['end_dates'] = est_dates + [time_stamps[-1]] + end_store
        df_store.append(df)

    excel_name = create_excel_file(f'{results_dir}/poos_analysis_{model_full_name}.xlsx')
    wb = openpyxl.Workbook()
    for df, h in zip(df_store, h_store):
        wb.create_sheet(f'h{h}')
        ws = wb[f'h{h}']
        print_df_to_excel(df=df, ws=ws)

    wb.save(excel_name)


def poos_model_evaluation(fl_master, ar_store, pca_store, xgb_stores, results_dir, blocked_dates, blocks,
                          first_est_date, est_dates, combinations=None):
    def xl_test(x):
        n = x.shape[0]
        u = x - x.mean()
        sigma = (x ** 2).mean()
        v = u ** 2 - sigma
        z = np.concatenate((u[None, :], v[None, :]), axis=0)
        ar1_model = sm.tsa.ARMA(u, (1, 0)).fit(disp=False)
        rho = abs(ar1_model.params[1])
        q = int(np.round((1.5 * n) ** (1 / 3) * (2 * rho / (1 - rho ** 2)) ** (2 / 3)))

        sum_uu = 0
        sum_uv = 0
        sum_vu = 0
        sum_vv = 0
        for h in range(1, q + 1):
            g_uu = 1 / n * np.sum(u[h:] * u[:-h])
            g_uv = 1 / n * np.sum(u[h:] * v[:-h])
            g_vu = 1 / n * np.sum(v[h:] * u[:-h])
            g_vv = 1 / n * np.sum(v[h:] * v[:-h])
            sum_uu += (1 - h / (q + 1)) * g_uu
            sum_uv += (1 - h / (q + 1)) * g_uv
            sum_vu += (1 - h / (q + 1)) * g_vu
            sum_vv += (1 - h / (q + 1)) * g_vv

        g_uu_0 = 1 / n * np.sum(u * u)
        g_uv_0 = 1 / n * np.sum(u * v)
        g_vv_0 = 1 / n * np.sum(v * v)

        w_u = g_uu_0 + 2 * sum_uu
        w_v = g_vv_0 + 2 * sum_vv
        w_uv = g_uv_0 + sum_uv + sum_vu
        if w_uv < 0:
            print('w_uv is negative')
        omega = np.array([[w_u, abs(w_uv) ** 0.5], [abs(w_uv) ** 0.5, w_v]])
        omega_neghalf = fractional_matrix_power(omega, -0.5)

        sequence = np.empty_like(z)
        z_cumsum = np.cumsum(z, axis=1)
        for i in range(z.shape[1]):
            sequence[:, i] = omega_neghalf @ z_cumsum[:, i]

        c = 1 / n ** 0.5 * np.max(np.linalg.norm(sequence, ord=1, axis=0))
        if c <= 1.89:
            p = 0.1
        elif c <= 2.07:
            p = 0.05
        elif c <= 2.4:
            p = 0.01
        else:
            p = 0
        return p

    def remove_lr_blocks(df, idx, v, slice_mode='number'):
        if slice_mode == 'number':
            return df.drop(df.index[idx - v:idx + v + 1])
        elif slice_mode == 'index':
            idx = np.where(df.index == idx)[0][0]
            return df.drop(df.index[idx - v:idx + v + 1])
        else:
            raise KeyError('slice_mode should be number or index.')

    def compute_mcs(e2_df):
        mcs = MCS(e2_df.copy(), size=0.1)
        try:
            mcs.compute()
            return mcs.pvalues.T.squeeze()
        except IndexError:
            # Means it is the 1970:1 block where ssm and ssm full is exactly the same.
            # So need to remove ssm full
            blank = pd.Series()
            if (e2_df['llc_full_ehat^2'] == e2_df['llc_ehat^2']).all():
                e2_df = e2_df.drop('llcfull_ehat^2', axis=1)
                blank['llc_full_ehat^2'] = 0
            if (e2_df['hparam_ehat^2'] == e2_df['max_iter_ehat^2']).all():
                e2_df = e2_df.drop('max_iter_ehat^2', axis=1)
                blank['max_iter_ehat^2'] = 0
            mcs = MCS(e2_df.copy(), size=0.1)
            mcs.compute()
            series = mcs.pvalues.T.squeeze()
            series = pd.concat((series, blank))
            return series

    def compute_spa(e2_df):
        series = {}
        for col in e2_df.columns.values:
            spa = SPA(e2_df[col], e2_df.drop(col, axis=1))
            spa.seed(42)
            spa.compute()
            series[col] = spa.pvalues['consistent']
        return pd.Series(series)

    def plotting_mcs(df, save_dir):
        df = df.copy()
        df = df[[x for y in ['ar_ar', 'pca', 'xgba(rh)_rw_', 'xgba(rfcv)_rw_', 'rf(rh)'] for x in df.columns if y in x]]
        df.columns = ['AR', 'PCA', 'XGBA(rh)_rw', 'XGBA(rfcv)_rw', 'RF(rh)']
        index = df.index.values
        blocks = list(range(10)) * 5
        index = [[x.partition('_')[0], block + 1] for block, x in zip(blocks, index)]
        df.index = pd.MultiIndex.from_arrays(np.array(index).T.tolist())

        f, a = plt.subplots(3, 2, constrained_layout=True)
        df.xs('h1').plot(ax=a[0, 0], legend=False)
        df.xs('h3').plot(ax=a[0, 1], legend=False)
        df.xs('h6').plot(ax=a[1, 0], legend=False)
        df.xs('h12').plot(ax=a[1, 1], legend=False)
        df.xs('h24').plot(ax=a[2, 0], legend=False)
        a[0, 0].set_title('h = 1')
        a[0, 1].set_title('h = 3')
        a[1, 0].set_title('h = 6')
        a[1, 1].set_title('h = 12')
        a[2, 0].set_title('h = 24')
        f.delaxes(a[2, 1])

        axLine, axLabel = f.axes[0].get_legend_handles_labels()
        lines = axLine
        labels = axLabel
        f.legend(lines, labels, loc='lower right')
        plt.savefig(save_dir, bbox_inches='tight')
        plt.close()

        f, a = plt.subplots(3, 2, constrained_layout=True)
        df.xs('h1').plot.bar(ax=a[0, 0], legend=False)
        df.xs('h3').plot.bar(ax=a[0, 1], legend=False)
        df.xs('h6').plot.bar(ax=a[1, 0], legend=False)
        df.xs('h12').plot.bar(ax=a[1, 1], legend=False)
        df.xs('h24').plot.bar(ax=a[2, 0], legend=False)
        a[0, 0].set_title('h = 1')
        a[0, 1].set_title('h = 3')
        a[1, 0].set_title('h = 6')
        a[1, 1].set_title('h = 12')
        a[2, 0].set_title('h = 24')
        f.delaxes(a[2, 1])

        axLine, axLabel = f.axes[0].get_legend_handles_labels()
        lines = axLine
        labels = axLabel
        f.legend(lines, labels, loc='lower right')
        plt.savefig(f'{save_dir.partition(".png")[0]}_bar.png', bbox_inches='tight')
        plt.close()

    def get_keyword_df_from_data(data, keyword, column_header=None):
        df = data['data_df'].copy()
        df = df[[x for x in df.columns if keyword in x]]
        if column_header:
            df.columns = [f'{column_header}_{x}' for x in df.columns]
        return df

    idx = fl_master.time_stamp.index(first_est_date)
    ts = fl_master.time_stamp[idx:]

    h_store = 1, 3, 6, 12, 24
    results_store = {}
    # selected_store = [[('drop',None),('drop',None),('drop',None),('drop',None),],
    #                  [('drop','oracle'),('drop','oracle'),('drop','oracle'),('drop','oracle'),],
    #                  [('drop','oracle'),('keep',None),('keep',None),('keep',None),],
    #                  [('keep',['_rw_','_ll*ln_','_rw+ll*ln_',]),('keep',None),('keep',None),('keep',None),],
    #                  [('keep','_rw_'),('keep',None),('keep',None),('keep',None),],]
    # name_store = ['all', 'all-oracle', 'xgbarh-oracle', 'xgbarh(selected)', 'xgbarh(rw)']
    selected_store = [
        [('keep', ['_rw_', '_hparam_', '_ll_e', '_ll*ln_', '_rw+ll*ln_', ]),
         ('keep', ['_rw_', '_hparam_', '_ll_e', '_ll*ln_', '_rw+ll*ln_', ]), ('drop', None), ('drop', None), ],
        # [('drop', None), ('drop', None), ('drop', None), ('drop', None), ],
        # [('drop', 'oracle'), ('drop', 'oracle'), ('drop', 'oracle'), ('drop', 'oracle'), ],
        # [('drop', 'oracle'), ('keep', None), ('keep', None), ('keep', None), ],
        # [('keep', '_rw_'), ('keep', None), ('keep', None), ('keep', None), ],
    ]
    name_store = ['xgba(selected)']  # 'all', 'all-oracle', 'xgbarh-oracle',  'xgbarh(rw)']
    xgb_names = ['xgba(rh)', 'xgba(rfcv)', 'rf(rh)', 'rf(rfcv)']
    mcs_store = collections.defaultdict(dict)
    spa_store = collections.defaultdict(dict)
    if blocks:
        est_dates = est_dates[1:]
        est_dates_idx = [0] + [ts.index(est) + 1 for est in est_dates] + [-1]

        for h, ar, pca, xgb_store in zip(h_store, ar_store, pca_store, xgb_stores):
            with open(ar, 'rb') as handle:
                ar_data = pickle.load(handle)
                ar_ehatdf = get_keyword_df_from_data(ar_data, keyword='_ehat', column_header='ar')
            with open(pca, 'rb') as handle:
                pca_data = pickle.load(handle)
                pca_ehatdf = get_keyword_df_from_data(pca_data, '_ehat', column_header='pca')
            xgb_ehatdf_store = []
            for xgb, name in zip(xgb_store, xgb_names):
                with open(xgb, 'rb') as handle:
                    xgb_data = pickle.load(handle)
                    combinations_store = []
                    data_df = get_keyword_df_from_data(xgb_data, '_ehat', column_header=name)
                    if combinations and 'xgb' in name:
                        for combi in combinations:
                            combi_name = '+'.join(combi) + '_ehat'
                            combinations_store.append(
                                pd.DataFrame(xgb_data['data_df'][[f'y_{h}']].values.squeeze() -
                                             (xgb_data['data_df'][[f'y_{h}']].values
                                              - data_df[[f'{name}_{x}_ehat' for x in combi]]).mean(axis=1),
                                             columns=[f'{name}_{combi_name}']))
                        xgb_ehatdf = pd.concat([data_df] + combinations_store, axis=1)
                    else:
                        xgb_ehatdf = get_keyword_df_from_data(xgb_data, '_ehat', column_header=name)
                xgb_ehatdf_store.append(xgb_ehatdf)

            model_data = pd.concat([pca_ehatdf] + xgb_ehatdf_store, axis=1)
            d_vectors = (ar_ehatdf['ar_ar_ehat'].values ** 2)[..., None] - \
                        model_data.values ** 2

            losses_data = pd.concat([ar_ehatdf, pca_ehatdf] + xgb_ehatdf_store, axis=1) ** 2
            losses_data.columns = [f'{x}^2' for x in losses_data.columns]

            for start, end in zip(est_dates_idx[:-1], est_dates_idx[1:]):
                p_values = np.apply_along_axis(lambda x: sm.tsa.stattools.kpss(x),
                                               axis=0, arr=d_vectors[start:end])[1, :]
                # p_xl_values = np.apply_along_axis(lambda x: xl_test(x),
                #                                  axis=0, arr=d_vectors[start:end])
                results_store[f'h{h}_{ts[start]}~{ts[end]}'] = np.concatenate((p_values, p_values))

                # series = compute_mcs(e2_df=losses_data.iloc[start:end].copy(), drop_cols=None)
                # series_no_oracle = compute_mcs(e2_df=losses_data.iloc[start:end].copy(), drop_cols=['oracle_ehat^2'])
                # mcs_store[f'h{h}_{ts[start]}~{ts[end]}'] = series
                # mcs_store_no_oracle[f'h{h}_{ts[start]}~{ts[end]}'] = series_no_oracle

                for name, selection in zip(name_store, selected_store):
                    df = losses_data.copy()
                    for select, xgb_name in zip(selection, xgb_names):
                        if isinstance(select[1], str):
                            variants = [select[1]]
                        else:
                            variants = select[1]
                        if not variants:
                            variants = []
                        if select[0] == 'drop':
                            column_drop = [x for x in losses_data.columns if
                                           xgb_name in x and any([y in x for y in variants])]
                        else:
                            column_drop = [x for x in losses_data.columns if
                                           xgb_name in x and not any([y in x for y in variants])]
                        df = df.drop(column_drop, axis=1)

                    spa_store[name][f'h{h}_{ts[start]}~{ts[end]}'] = compute_spa(
                        df.iloc[start:end])

                    mcs_store[name][f'h{h}_{ts[start]}~{ts[end]}'] = compute_mcs(
                        df.iloc[start:end])

    else:
        for h, ar, pca, xgb in zip(h_store, ar_store, pca_store, xgb_store):
            with open(ar, 'rb') as handle:
                ar_data = pickle.load(handle)

            with open(pca, 'rb') as handle:
                pca_data = pickle.load(handle)

            with open(xgb, 'rb') as handle:
                xgb_data = pickle.load(handle)

            model_data = pd.concat([pca_data['data_df'], xgb_data['data_df']], axis=1)

            d_vectors = (ar_data['data_df']['ar_ehat'].values ** 2)[..., None] - \
                        model_data[[x for x in model_data.columns.values if '_ehat' in x]].values ** 2

            p_values = np.apply_along_axis(lambda x: sm.tsa.stattools.kpss(x),
                                           axis=0, arr=d_vectors)[1, :]

            p_xl_values = np.apply_along_axis(lambda x: xl_test(x),
                                              axis=0, arr=d_vectors)

            # Plotting
            plt.close()
            d_vectors = pd.DataFrame(d_vectors, columns=[x.replace('_ehat', '_d') for x in model_data.columns.values if
                                                         '_ehat' in x], index=ts)
            d_vectors.plot(subplots=True, layout=(3, 2))
            plt.savefig(f'{results_dir}/h{h}_d_plot.png')

            v = 12
            for dates in blocked_dates:
                d_vectors = remove_lr_blocks(d_vectors, idx=dates, v=v, slice_mode='index')

            d_vectors.plot(subplots=True, layout=(3, 2))
            plt.savefig(f'{results_dir}/h{h}_d_removed_plot.png')

            d_vectors = d_vectors.values

            p_values_adj = np.apply_along_axis(lambda x: sm.tsa.stattools.kpss(x),
                                               axis=0, arr=d_vectors)[1, :]

            p_xl_values_adj = np.apply_along_axis(lambda x: xl_test(x),
                                                  axis=0, arr=d_vectors)

            results_store[f'h{h}'] = [f'{x:.3f} / {y:.3f}' for x, y in zip(p_values, p_values_adj)] + [
                f'{x:.3f} / {y:.3f}' for x, y in zip(p_xl_values, p_xl_values_adj)]

    results_df = pd.DataFrame.from_dict(results_store, orient='index',
                                        columns=[x.replace('_ehat', type) for type in ['_KPSS', '_XL'] for x in
                                                 model_data.columns.values if '_ehat' in x])

    wb = openpyxl.Workbook()
    ws = wb[wb.sheetnames[-1]]
    print_df_to_excel(df=results_df, ws=ws)

    writer = pd.ExcelWriter(f'{results_dir}/test_summary.xlsx', engine='openpyxl')
    writer.book = wb

    if blocks:
        for (k1, v1), (k2, v2) in zip(mcs_store.items(), spa_store.items()):
            results_df1 = pd.DataFrame(v1).T
            plotting_mcs(results_df1, save_dir=f'{results_dir}/MCS_{k1}.png')
            mean = results_df1.mean()
            count10 = results_df1[results_df1 < 0.1].count().astype(str)
            results_df1 = results_df1.round(2).astype(str)
            results_df1.loc['mean'] = mean.round(2).astype(str)
            results_df1.loc['<10%'] = count10

            v1 = results_df1
            index = [[x.partition('_')[0], x.partition('_')[-1]] for x in v1.columns]
            v1.columns = pd.MultiIndex.from_arrays(np.array(index).T.tolist())
            v1.to_excel(writer, sheet_name=f'MCS_{k1}')

            results_df2 = pd.DataFrame(v2).T
            plotting_mcs(results_df2, save_dir=f'{results_dir}/SPA_{k2}.png')
            mean = results_df2.mean()
            count10 = results_df2[results_df2 < 0.1].count().astype(str)
            results_df2 = results_df2.round(2).astype(str)
            results_df2.loc['mean'] = mean.round(2).astype(str)
            results_df2.loc['<10%'] = count10

            v2 = results_df2
            index = [[x.partition('_')[0], x.partition('_')[-1]] for x in v2.columns]
            v2.columns = pd.MultiIndex.from_arrays(np.array(index).T.tolist())
            v2.to_excel(writer, sheet_name=f'SPA_{k2}')

            v1.iloc[:-1, :] = v1.iloc[:-1, :]
            v2.iloc[:-1, :] = v2.iloc[:-1, :]
            v3 = v1.astype(str) + ' / ' + v2.astype(str)
            v3.to_excel(writer, sheet_name=f'all_{k1}')

    writer.save()
    writer.close()


def combine_poos_excel_results(excel_store, results_dir, name_store, selected_xgba, expt_type='expt1'):
    set_matplotlib_style()

    def get_df_for_plotting(v, col_names, selected_col):
        v = v[[x for y in selected_col for x in v.columns if
               fnmatch.fnmatch(f'{x[0]}_{x[1]}', y)]]
        v.columns = col_names
        v = v.drop(v.index[-3])  # drop -3 to remove COVID only months
        v.index = v.index.values.tolist()[:-7] + ['Full less COVID', 'Full', 'Pre-Great Moderation',
                                                  'Great Moderation', 'Post-Great Moderation', 'Expansionary',
                                                  'Recessionary']
        return v

    data_store = {f'h{x}': [] for x in [1, 3, 6, 12, 24]}
    for idx, (excel_dir, name) in enumerate(zip(excel_store, name_store)):
        xls = pd.ExcelFile(excel_dir)
        for (k, v) in data_store.items():
            df = pd.read_excel(xls, k, index_col=0)
            if idx == 0:
                v.append(df[['start_dates', 'end_dates']])
            df.drop(['start_dates', 'end_dates'], inplace=True, axis=1)
            df.columns = [f'{name}_{x}' for x in df.columns]
            v.append(df)

    results_dir = create_results_directory(results_dir)
    excel_name = create_excel_file(f'{results_dir}/combined_poos_results.xlsx')
    wb = openpyxl.load_workbook(excel_name)
    for k, v in data_store.items():
        wb.create_sheet(k)
        ws = wb[k]
        v = pd.concat(v, axis=1)
        print_df_to_excel(df=v, ws=ws)
    wb.save(excel_name)
    new_index = [f'{x}~{y}' for x, y in zip(v['start_dates'], v['end_dates'])]
    book = openpyxl.load_workbook(excel_name)
    writer = pd.ExcelWriter(excel_name, engine='openpyxl')
    writer.book = book

    for k, v in data_store.items():
        # RMSE columns only
        v = pd.concat(v, axis=1)
        v = v[[x for x in v.columns if '_rmse' in x]]
        v = v.T
        index = [[x.partition('_')[0], x.partition('_')[-1]] for x in v.index]
        v.index = pd.MultiIndex.from_arrays(np.array(index).T.tolist())
        v = v.T
        v.index = new_index
        v.columns = pd.MultiIndex.from_tuples([(x, y.replace('_rmse', '')) for x, y in v.columns])
        v.to_excel(writer, sheet_name=f'{k}_rmse')

        # Selected xgba columns with rw, ar, pca, rf.
        v = v.drop([(model, variant) for model, variant in v.columns if
                    'xgba' in model and not variant in selected_xgba], axis=1)
        v.to_excel(writer, sheet_name=f'{k}_rmse_sel')

        # Relative to RW
        v_abs = v.copy()
        v[v.columns[1:]] = v[v.columns[1:]] / v[[v.columns[0]]].values
        v.to_excel(writer, sheet_name=f'{k}_rel_rmse_sel')

        if expt_type == 'poos':
            # Plotting relative RMSE against blocks
            # v = v[[x for x in v.columns if any(y in x for y in ['pca', 'rw', 'rf'])]]
            v = get_df_for_plotting(v, col_names=['AR', 'PCA', 'XGBA(rh)_rw', 'RF(rh)'],
                                    selected_col=['ar_ar', 'pca_pca', 'xgba(rh, *)_rw', 'rf(rh, *)_rf'])
            v_abs = get_df_for_plotting(v_abs, col_names=['RW'],
                                    selected_col=['rw_rw'])

            fig, ax1 = plt.subplots()
            v.plot.bar(ax=ax1)
            ax1.set_ylabel('Relative RMSE against RW', size=10)
            plt.axvspan(10 - 0.5, v.shape[0], facecolor='grey', alpha=0.2)
            if '1' in k and '2' not in k:
                plt.legend(loc=2)
            else:
                plt.legend('')

            ax2 = ax1.twinx()
            ax2.plot(v_abs.values, alpha=0.35, color='black')
            ax2.set_ylabel('RW RMSE', size=10)


            plt.savefig(f'{results_dir}/{k}_bar.png', bbox_inches='tight')
            plt.close()

    writer.save()
    writer.close()


class Shap_data:
    def __init__(self, shap_matrix, feature_names, ts_index=None):
        df = pd.DataFrame(shap_matrix, columns=feature_names).iloc[:, 1:].T  # Remove first column of constants
        # Convert to multi index. Split name and lags
        df.index = pd.MultiIndex.from_tuples(df.index.str.split('_L').tolist())
        self.feature_names = feature_names
        self.df = df.T  # Change back to multi-column
        self.grouped_df = self.df.abs().sum(level=0, axis=1)
        self.shap_abs = self.df.abs().sum(axis=0)
        self.grouped_shap_abs = self.grouped_df.sum(axis=0)
        if ts_index:
            self.df.index = ts_index
            self.grouped_df = ts_index

    def summary_plot(self, grouped=False, plot_type='bar'):
        if grouped:
            shap.summary_plot(self.grouped_df, feature_names=self.grouped_df.columns.values, plot_type=plot_type)
        else:
            shap.summary_plot(self.df, feature_names=self.df.columns.values, plot_type=plot_type)

    def add_feature_data(self, fl_master, fl, hparams, h, ts, expected_value, update_ts_only=False):
        if not update_ts_only:
            h_idx = np.where(np.array([1, 3, 6, 12, 24]) == h)[0][0]
            idx = np.where(np.array(fl_master.time_stamp) == ts)[0][0]
            ts_index = fl_master.time_stamp[idx - self.df.shape[0] + 1:idx + 1]
            z_matrix, y_vec = fl.prepare_data_matrix(fl_master.x[:idx + 1, :], fl_master.yo[:idx + 1, :],
                                                     fl_master.y[:idx + 1, [h_idx]],
                                                     h, int(hparams['m']), int(hparams['m'] * 2), 1)

            self.z_matrix = pd.DataFrame(z_matrix, columns=self.feature_names).drop('constant', axis=1).iloc[
                            -self.df.shape[0]:, :]
            self.z_matrix.index = ts_index
            self.y_vec = y_vec
            self.expected_value = expected_value
            self.df.index = ts_index
            self.grouped_df.index = ts_index
        else:
            idx = np.where(np.array(fl_master.time_stamp) == ts)[0][0]
            ts_index = fl_master.time_stamp[idx - self.df.shape[0] + 1:idx + 1]
            self.df.index = ts_index
            self.grouped_df.index = ts_index

    def add_feature_info(self, feature_info_df, h):
        '''
        df = {}
        for item in self.shap_abs.iteritems():
            info = feature_info_df[feature_info_df['Name'] == item[0][0]].squeeze().to_dict()
            if len(info['Group']) == 0:
                if item[0][0] == 'y':
                    info = {'Group': "AR", 'Name': 'AR', "Description": 'AR'}
                else:
                    raise KeyError(f'Feature name {item[0][0]} not found in feature info excel.')
            df[f'{item[0][0]}_L{item[0][1]}'] = {**info, **{'|SHAP|': item[1], 'L': item[0][1]}}

        df = pd.DataFrame.from_dict(df, orient='index')
        df['h'] = h
        self.feature_info_df = df
        '''

        multi_column = [
            (var, feature_info_df['Group'].loc[feature_info_df['Name'] == var].values[0]) if not var == 'y' else (
                var, 'AR') for var in self.grouped_df.columns]
        self.grouped_df.columns = pd.MultiIndex.from_tuples(multi_column)


def poos_shap(fl_master, fl, xgb_store, first_est_date, results_dir, feature_info_dir, est_dates=None,
              other_xgb_store=None):
    set_matplotlib_style()

    feature_info_df = pd.read_excel(feature_info_dir, index_col=0)

    excel_name = create_excel_file(f'{results_dir}/poos_shap.xlsx')
    wb = openpyxl.load_workbook(excel_name)

    idx = fl_master.time_stamp.index(first_est_date)
    ts = fl_master.time_stamp[idx:]
    info_df = {}
    # fi_df_store = []

    def get_top_columns_per_row(df, n_top):
        ranked_matrix = np.argsort(-df.values, axis=1)[:, :n_top]
        return pd.DataFrame(df.columns.values[ranked_matrix],
                            index=df.index,
                            columns=[f'Rank {idx + 1}' for idx in range(n_top)]), \
               pd.DataFrame(df.values[np.repeat(np.arange(df.shape[0])[:, None], n_top, axis=1), ranked_matrix],
                            index=df.index,
                            columns=[f'Rank {idx + 1}' for idx in range(n_top)])

    def print_df(sheet_name, df, h):
        wb.create_sheet(f'h{h}_{sheet_name}')
        ws = wb[f'h{h}_{sheet_name}']
        print_df_to_excel(ws=ws, df=df)

    def get_grouped_df(xgb_data, first_est_date, ts_index):
        poos_data_store = dict(zip(ts_index,
                                   [{'data': x,
                                     'hparam': blocks['hparams_df'].iloc[0, :],
                                     'feature_names':blocks['poos_data_store'][0]['feature_names']
                                     } for blocks in xgb_data for x in blocks['poos_data_store']]))
        grouped_df = []
        sd = Shap_data(poos_data_store[first_est_date]['data']['shap_values'].toarray(), poos_data_store[first_est_date]['feature_names'])
        sd.add_feature_data(fl_master, fl, hparams=poos_data_store[first_est_date]['hparam'], h=h,
                            ts=first_est_date, update_ts_only=True,
                            expected_value=poos_data_store[first_est_date]['data']['expected_value'])
        sd.add_feature_info(feature_info_df, h=h)
        grouped_df.append(sd.grouped_df.copy())
        for date, v in poos_data_store.items():
            if date == first_est_date:
                continue
            sd = Shap_data(v['data']['shap_values'].toarray()[[-1],:], poos_data_store[date]['feature_names'])
            sd.add_feature_data(fl_master, fl, hparams=v['hparam'], h=h,
                                ts=date, update_ts_only=True,
                                expected_value=v['data']['expected_value'])
            sd.add_feature_info(feature_info_df, h=h)
            grouped_df.append(sd.grouped_df.iloc[[-1], :].copy())

        return pd.concat(grouped_df, axis=0)

    def get_grouped_df_for_poos(xgb_data):
        poos_data_store = dict(zip(ts, [x for blocks in xgb_data for x in blocks['poos_data_store']]))
        est_data_store = {k: v for k, v in poos_data_store.items() if k in est_dates}
        hparam_store = [data['hparams_df'].iloc[0, :] for data in xgb_data + [xgb_data[-1]]]
        grouped_df = []
        est_previous = None
        for (est, data), hparam in zip(est_data_store.items(), hparam_store):
            try:
                feature_names = data['feature_names']
            except KeyError:
                # feature name is not in dict for the 2020:1 data dict. But the feature name would had been
                # defined in 2015:1 which will mean it is ok
                pass
            sd = Shap_data(data['shap_values'].toarray(), feature_names)
            sd.add_feature_data(fl_master, fl, hparams=hparam, h=h, ts=est,
                                expected_value=data['expected_value'], update_ts_only=True)
            sd.add_feature_info(feature_info_df, h=h)
            if est_previous:
                temp_df = sd.grouped_df.iloc[sd.grouped_df.index.get_loc(est_previous) + 1:, :]
            else:
                temp_df = sd.grouped_df
            index = [(idx, est) for idx in temp_df.index]
            temp_df.index = pd.MultiIndex.from_tuples(index)
            grouped_df.append(temp_df.copy())
            est_previous = est

        return pd.concat(grouped_df, axis=0)

    def norm_grouped_df_plotting(grouped_df, name='finalmodel', expt_type='expt1'):
        if expt_type == 'poos':
            grouped_df.index = grouped_df.index.get_level_values(0)
        norm_grouped_df = grouped_df.div(grouped_df.sum(axis=1), axis=0)

        # Scatter plot for Pre-GM and GM
        ts_gm = norm_grouped_df.index.get_loc('1984:4')
        ts_postgm = norm_grouped_df.index.get_loc('2007:1')
        temp_df = pd.concat([norm_grouped_df.iloc[:ts_gm,:].mean(axis=0), norm_grouped_df.iloc[ts_gm:ts_postgm,:].mean(axis=0)], axis=1).reset_index()
        temp_df.columns = ['Feature Name', 'Group', 'Pre-Great Moderation', 'Great Moderation']
        ax=sns.scatterplot(data=temp_df, x='Pre-Great Moderation', y='Great Moderation', hue='Group', s=20)
        handles, labels = ax.get_legend_handles_labels()
        ax.legend(handles=handles[1:], labels=labels[1:], fontsize=8.5)
        sns.despine()
        plt.savefig(f'{results_dir}/h{h}_{name}_scatterplot.png', bbox_inches='tight')
        plt.close()

        # Stacked area plot for SHAP vs time with hue=Groups
        norm_grouped_df.groupby(level=1, axis=1).sum().rolling(12, min_periods=1).mean().plot.area().legend(
            loc='center left', bbox_to_anchor=(1, 0.5))
        fig = plt.gcf()
        fig.set_size_inches(10, 4)
        sns.despine()
        plt.xlim(xmin=12)
        plt.ylabel('Normalized Grouped |SHAP|')
        plt.savefig(f'{results_dir}/h{h}_{name}_stackedplot.png', bbox_inches='tight')
        plt.close()

        # Joyplot for Distribution of SHAP values across time averaged in 11 blocks
        temp_df = norm_grouped_df.reset_index()
        temp_df = temp_df.groupby(temp_df.index // 12).mean()
        temp_index = norm_grouped_df.index[::12]
        temp_df.index = temp_index
        temp_df = temp_df.reset_index().rename(
            columns={'index': 'Date'}).melt(id_vars='Date', var_name=['Var', 'Group'],
                                            value_name='|SHAP|')
        labels = [y if int(y.partition(':')[0]) % 10 == 0 else None for y in list(temp_index)]
        fig, axes = joypy.joyplot(temp_df, by="Date", column="|SHAP|", labels=labels,
                                  range_style='own', x_range=[0, 0.1], fade=True,
                                  grid="y", linewidth=1, legend=False, figsize=(6, 6),
                                  colormap=cm.autumn_r)
        plt.savefig(f'{results_dir}/h{h}_{name}_ridgelineplot.png', bbox_inches='tight')
        plt.close()


    for xgb_dir, h in zip(xgb_store, [1, 3, 6, 12, 24]):
        with open(xgb_dir, 'rb') as handle:
            xgb_data = pickle.load(handle)

        if est_dates:
            # [f'{x}:1' for x in range(1970, 2020, 5)] + ['2020:6']
            grouped_df = get_grouped_df(xgb_data=xgb_data, first_est_date=first_est_date, ts_index=ts)
            norm_grouped_df_plotting(grouped_df.copy(), 's42', expt_type='poos')
            if other_xgb_store:
                grouped_df_store = [grouped_df.copy()]
                for other_xgb_dict in other_xgb_store:
                    with open(other_xgb_dict[h], 'rb') as handle:
                        other_xgb_data = pickle.load(handle)
                    grouped_df_store.append(get_grouped_df(other_xgb_data, first_est_date=first_est_date, ts_index=ts))
                grouped_df = pd.concat(grouped_df_store, axis=0).groupby(axis=0, level=0).mean()
                norm_grouped_df_plotting(grouped_df, 'averaged', expt_type='poos')

        shap_abs = []
        grouped_shap_abs = []

        for block in xgb_data:
            feature_names = block['poos_data_store'][0]['feature_names']
            for step in block['poos_data_store']:
                shap_data = Shap_data(shap_matrix=step['shap_values'].toarray(), feature_names=feature_names)
                shap_abs.append(shap_data.shap_abs)
                grouped_shap_abs.append(shap_data.grouped_shap_abs)

        shap_abs = pd.concat(shap_abs, axis=1).T
        shap_abs.index = ts
        gshap_abs = pd.concat(grouped_shap_abs, axis=1).T
        gshap_abs.index = ts



        top_shap_names, top_shap_values = get_top_columns_per_row(df=shap_abs, n_top=10)
        top_gshap_names, top_gshap_values = get_top_columns_per_row(df=gshap_abs, n_top=10)

        # gshap_abs_norm = gshap_abs.div(gshap_abs.max(axis=1), axis=0)
        # gshap_abs.plot(y=np.unique(top_gshap_names.values))

        print_df('shap_names', df=top_shap_names.applymap(str), h=h)
        print_df('shap_values', df=top_shap_values, h=h)
        print_df('gshap_names', df=top_gshap_names.applymap(str), h=h)
        print_df('gshap_values', df=top_gshap_values, h=h)

        # SHAP plots
        poos_data_store = dict(zip(ts,
                                   [{'data': x, 'hparam': blocks['hparams_df'].iloc[0, :]} for blocks in xgb_data for x
                                    in blocks['poos_data_store']]))
        for date in ['2020:6']:
            sd = Shap_data(poos_data_store[date]['data']['shap_values'].toarray(), feature_names)
            sd.add_feature_data(fl_master, fl, hparams=xgb_data[-1]['hparams_df'].iloc[0, :], h=h, ts=date,
                                expected_value=poos_data_store[date]['data']['expected_value'])

            shap.summary_plot(sd.df.values, sd.z_matrix, max_display=7, show=False)
            plt.savefig(f'{results_dir}/h{h}_{date.replace(":", "_")}.png', bbox_inches='tight')
            plt.close()

            z_scores = np.abs(zscore(sd.df))
            filtered_values = sd.df.copy().values
            filtered_values[z_scores > 3] = 0
            shap.summary_plot(filtered_values, sd.z_matrix, max_display=7, show=False)
            plt.savefig(f'{results_dir}/h{h}_{date.replace(":", "_")}_removedoutlier.png', bbox_inches='tight')
            plt.close()

            info_df[f'h{h}_{date}'] = {'top7_frac': sd.shap_abs.nlargest(7).sum() / sd.shap_abs.sum(),
                                       'outliers': np.count_nonzero(z_scores > 3)}

            shap.force_plot(sd.expected_value, sd.df.values[-1, :], sd.z_matrix.iloc[-1, :].round(3), matplotlib=True,
                            show=False)
            plt.savefig(f'{results_dir}/h{h}_{date.replace(":", "_")}_forceplot.png', bbox_inches='tight')
            plt.close()

            if date == '2020:6' and not est_dates:
                sd.add_feature_info(feature_info_df, h=h)
                # fi_df_store.append(sd.feature_info_df)

                grouped_df = get_grouped_df(xgb_data, first_est_date, fl_master.time_stamp[idx:])
                norm_grouped_df_plotting(grouped_df, 's42')

                if other_xgb_store:
                    grouped_df_store = [grouped_df.copy()]
                    for other_xgb_dict in other_xgb_store:
                        with open(other_xgb_dict[h], 'rb') as handle:
                            data = pickle.load(handle)
                        grouped_df_store.append(get_grouped_df(data, first_est_date, fl_master.time_stamp[idx:]))
                    grouped_df = pd.concat(grouped_df_store, axis=0).groupby(axis=0, level=0).mean()
                    norm_grouped_df_plotting(grouped_df=grouped_df, name='s42,100,200,300')

    '''
    if not est_dates:
        fi_df = pd.concat(fi_df_store, axis=0)
        fi_df['|SHAP|'] = fi_df[['|SHAP|', 'h']].groupby('h').transform(lambda x: x / (x.sum()))

        sns.barplot(x='h', y='|SHAP|', hue='Group', data=fi_df.groupby(['h', 'Group']).sum().reset_index())
        plt.legend(ncol=2)
        plt.savefig(f'{results_dir}/group_SHAP.png')

        temp = fi_df.groupby('h').apply(lambda x: x.nlargest(7, columns='|SHAP|')).reset_index(1)
        temp.index = list(np.arange(1, 8)) * 5
        ws = wb['Sheet']
        print_df_to_excel(df=temp, ws=ws)
    '''

    df = pd.DataFrame.from_dict(info_df).T
    wb.create_sheet('info')
    ws = wb['info']
    print_df_to_excel(ws=ws, df=df)
    wb.save(excel_name)
    wb.close()
