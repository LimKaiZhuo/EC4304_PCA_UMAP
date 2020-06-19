import numpy as np
import pandas as pd
import math, random
import cvxpy as cp
import openpyxl
import statsmodels.api as sm
from openpyxl.utils.dataframe import dataframe_to_rows

from own_package.postprocess import Postdata
from own_package.others import create_excel_file, print_df_to_excel, print_array_to_excel


def decomp_combi(var_name, numel, subgroup_size):
    results_dir = './results/{} Done'.format(var_name)
    post = Postdata(results_dir=results_dir, var_name=var_name, calculations=False, star=True)
    all_h_y_hat = [np.array(ar.tolist() + pca.tolist() + umap.tolist()) for ar, pca, umap in
                          zip(post.testset_AR_y_hat, post.testset_PCA_y_hat, post.testset_UMAP_y_hat)]
    model_count = [single_all_y_hat.shape[0] for single_all_y_hat in all_h_y_hat]
    if any(subgroup_size>= np.array(model_count)):
        raise ValueError('subgroup_size given is {} which is >= model_count value of {}.'
                         ' Choose a smaller subgroup_size'.format(subgroup_size, model_count))

    excel_dir = create_excel_file('./results/{} Done/decomp_combi.xlsx'.format(var_name))
    wb = openpyxl.load_workbook(excel_dir)

    selections = [random.sample(list(range(model_count[0])), k=subgroup_size) for _ in range(numel)]
    all_h_p_y_hat = []
    all_h_rmse = []
    for single_all_y_hat, single_y, h_label in zip(all_h_y_hat, post.testset_AR_y, post.hsteps):
        # perform sub selection for each h step ahead
        sub_y_hat_store = np.array([single_all_y_hat[selection, :] for selection in selections])
        sub_y_mean_hat = np.mean(sub_y_hat_store, axis=1)
        sub_y_invvar_hat = np.reciprocal(np.var(sub_y_hat_store, axis=1))
        total_weights = np.sum(sub_y_invvar_hat, axis=0)
        p_y = np.sum((1/total_weights * sub_y_mean_hat * sub_y_invvar_hat), axis=0)
        all_h_p_y_hat.append(p_y)
        all_h_rmse.append(np.sqrt(np.average(np.square(p_y-single_y))))
        wb.create_sheet('h={}'.format(h_label))
        ws = wb[wb.sheetnames[-1]]

        ws.cell(1, 1).value = 'numel'
        ws.cell(1, 2).value = numel
        ws.cell(1, 3).value = 'subgroup_size'
        ws.cell(1, 4).value = subgroup_size
        ws.cell(2,2).value = 'rmse'
        print_array_to_excel(array=single_y, first_cell=(3,3), ws=ws, axis=1)
        ws.cell(3,2).value = ''
        ws.cell(4,2).value = all_h_rmse[-1]
        print_array_to_excel(array=p_y, first_cell=(4,3), ws=ws, axis=1)

    wb.save(excel_dir)


